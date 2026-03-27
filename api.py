"""
Solvist API v5.0.0 — Opportunity Intelligence Layer (V1 Blueprint)
Replaces generic CRM endpoints with:
- GET /dashboard
- GET /activation
- GET /insights
- POST /tracking
- POST /engine/score-all (Monthly Scoring Job)
"""

# from __future__ import annotations (Removed to fix Pydantic ForwardRef issues with UploadFile)

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".local_deps"))

import uuid
import math
import logging
import secrets
import re
import json
import csv
import time
from threading import Lock
from textwrap import dedent
from datetime import datetime, timezone, date, timedelta
from typing import Annotated, Dict, List, Literal, Optional
from enum import Enum
from urllib.parse import urlparse, parse_qs
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import URLError, HTTPError

from fastapi import FastAPI, Depends, HTTPException, Request, Response, status, File, UploadFile, Form, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from slowapi import Limiter
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from slowapi.util import get_remote_address
from supabase import create_client, Client
from dotenv import load_dotenv

import io
from fpdf import FPDF
import os
from dotenv import load_dotenv

from apscheduler.schedulers.background import BackgroundScheduler
import pandas as pd
from openpyxl import load_workbook
from scoring.engine import (
    compute_opportunity_score,
    OPP_BATTERY_UPGRADE,
    OPP_INVERTER_REPLACEMENT,
    OPP_SYSTEM_EXPANSION,
    OPP_DISPLAY_NAMES,
    OPP_EV_CHARGER,
    OPP_MAINTENANCE,
    OPP_INDUSTRIAL_BATTERY
)
from config import (
    SUPABASE_URL,
    SUPABASE_SERVICE_ROLE_KEY,
    STRIPE_SECRET_KEY,
    RESEND_API_KEY,
    ENVIRONMENT,
    ALLOWED_ORIGINS,
)
import resend
import stripe
import psycopg2
from db import get_db_connection

def verify_supabase_token(token: str) -> dict:
    """Validate Supabase JWT against Supabase Auth API using service key."""
    if not admin_client:
        raise HTTPException(status_code=500, detail="Internal server error")
    try:
        user_res = admin_client.auth.get_user(token)
        user = getattr(user_res, "user", None)
        if not user or not getattr(user, "id", None):
            raise HTTPException(status_code=401, detail="Unauthorized")
        return {"sub": str(user.id), "email": getattr(user, "email", None)}
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=401, detail="Unauthorized")


if STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY
if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# ─── Logging & Env ─────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("solvist")

# FIX #3: admin_client initialized lazily — set in lifespan startup to avoid blocking module import
admin_client: Client = None  # type: ignore

OPPORTUNITY_DISPLAY_ES: Dict[str, str] = {
    OPP_BATTERY_UPGRADE: "Mejora con Batería",
    OPP_INVERTER_REPLACEMENT: "Sustitución de Inversor",
    OPP_SYSTEM_EXPANSION: "Ampliación del Sistema",
    OPP_EV_CHARGER: "Instalación de Cargador EV",
    OPP_MAINTENANCE: "Contrato de Mantenimiento",
    OPP_INDUSTRIAL_BATTERY: "Batería Industrial",
}

PIPELINE_STATUS_ES: Dict[str, str] = {
    "New": "Nuevo",
    "Contacted": "Contactado",
    "Proposal": "Propuesta",
    "Closed": "Cerrado",
    "Lost": "Perdido",
}

SALES_ACTION_DISPLAY_ES: Dict[str, str] = {
    "called": "Llamada realizada",
    "email_sent": "Email comercial enviado",
    "proposal_sent": "Propuesta enviada",
}

COUNTRY_TO_CURRENCY_FALLBACK: Dict[str, str] = {
    "PR": "USD",
    "US": "USD",
    "MX": "MXN",
    "CO": "COP",
    "CL": "CLP",
    "ES": "EUR",
}

IMPORT_MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
SCORING_BATCH_SIZE = 500
REQUEST_ID_HEADER = "X-Request-Id"
DEPLOY_VERSION = "v1.0.0"
LAST_IMPORT_STATUS: Dict[str, float] = {
    "last_import_rows": 0,
    "last_import_duration": 0.0,
    "last_import_opportunities": 0,
}


# ─── Models & Enums ───────────────────────────────────────────────────────────
class CurrentUserContext(BaseModel):
    id: str
    email: Optional[str] = None
    company_id: Optional[str] = None

    class Config:
        frozen = True


class TenantContext(BaseModel):
    user_id: str
    company_id: str
    jwt: str
    installation_limit: int

    class Config:
        frozen = True


def _auth_log(
    level: int,
    event: str,
    request: Optional[Request] = None,
    user_id: Optional[str] = None,
) -> None:
    request_id = getattr(getattr(request, "state", None), "request_id", "")
    endpoint = str(request.url.path) if request else ""
    logger.log(
        level,
        "AUTH event=%s request_id=%s user_id=%s endpoint=%s",
        event,
        request_id,
        user_id or "",
        endpoint,
    )


def _normalize_jwt_sub(jwt_sub: str) -> str:
    try:
        return str(uuid.UUID(str(jwt_sub)))
    except Exception:
        raise HTTPException(status_code=401, detail="Unauthorized")


def _is_service_role_token(token: str) -> bool:
    service_key = _normalize_secret(SUPABASE_SERVICE_ROLE_KEY)
    return bool(service_key) and secrets.compare_digest(token, service_key)


def _get_or_create_public_user(
    jwt_sub: str,
    jwt_email: Optional[str],
    request: Optional[Request] = None,
) -> Dict:
    user_id = _normalize_jwt_sub(jwt_sub)
    users_table = admin_client.schema("public").table("users")

    user: Optional[Dict] = None
    try:
        user_res = users_table.select("*").eq("id", user_id).limit(1).execute()
        user = (user_res.data or [None])[0]
    except Exception:
        _auth_log(logging.ERROR, "public_user_lookup_failed", request=request, user_id=user_id)
        raise HTTPException(status_code=500, detail="Internal server error")

    insert_payload: Dict[str, str] = {"id": user_id}
    if jwt_email:
        insert_payload["email"] = jwt_email

    if not user:
        try:
            created = users_table.insert(insert_payload).execute()
            user = (created.data or [None])[0]
            _auth_log(logging.INFO, "public_user_autocreated", request=request, user_id=user_id)
        except Exception:
            if "email" in insert_payload:
                _auth_log(
                    logging.WARNING,
                    "public_user_autocreate_retry_without_email",
                    request=request,
                    user_id=user_id,
                )
                try:
                    created = users_table.insert({"id": user_id}).execute()
                    user = (created.data or [None])[0]
                    _auth_log(logging.INFO, "public_user_autocreated", request=request, user_id=user_id)
                except Exception:
                    _auth_log(logging.ERROR, "public_user_autocreate_failed", request=request, user_id=user_id)
                    raise HTTPException(status_code=500, detail="Internal server error")
            else:
                _auth_log(logging.ERROR, "public_user_autocreate_failed", request=request, user_id=user_id)
                raise HTTPException(status_code=500, detail="Internal server error")

    if not user:
        try:
            user_res = users_table.select("*").eq("id", user_id).limit(1).execute()
            user = (user_res.data or [None])[0]
        except Exception:
            _auth_log(logging.ERROR, "public_user_lookup_failed", request=request, user_id=user_id)
            raise HTTPException(status_code=500, detail="Internal server error")

    if not user:
        _auth_log(logging.ERROR, "public_user_resolution_failed", request=request, user_id=user_id)
        raise HTTPException(status_code=500, detail="Internal server error")

    return user


def _resolve_current_user(token: str, request: Request) -> CurrentUserContext:
    if _is_service_role_token(token):
        _auth_log(logging.WARNING, "service_role_token_rejected", request=request)
        raise HTTPException(status_code=403, detail="Forbidden")

    payload = verify_supabase_token(token)
    jwt_sub = payload.get("sub")
    if not jwt_sub:
        _auth_log(logging.WARNING, "jwt_missing_sub", request=request)
        raise HTTPException(status_code=401, detail="Unauthorized")

    jwt_email = payload.get("email")
    user = _get_or_create_public_user(jwt_sub, jwt_email, request=request)

    user_id = str(user.get("id") or _normalize_jwt_sub(jwt_sub))
    email = user.get("email") or jwt_email
    company_id = user.get("company_id")
    current_user = CurrentUserContext(
        id=user_id,
        email=str(email) if email else None,
        company_id=str(company_id) if company_id else None,
    )
    _auth_log(logging.INFO, "authenticated_user", request=request, user_id=current_user.id)
    return current_user


def _build_tenant_context(current_user: CurrentUserContext, token: str) -> TenantContext:
    if not current_user.company_id:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        res_comp = (
            admin_client.table("companies")
            .select("installation_limit")
            .eq("id", current_user.company_id)
            .single()
            .execute()
        )
        max_inst = res_comp.data.get("installation_limit") or 500
    except Exception:
        max_inst = 500

    return TenantContext(
        user_id=current_user.id,
        company_id=current_user.company_id,
        jwt=token,
        installation_limit=max_inst,
    )


# ─── Auth Dependency ───────────────────────────────────────────────────────────
bearer_scheme = HTTPBearer(auto_error=False)

async def get_current_user(
    request: Request,
    credentials: Annotated[Optional[HTTPAuthorizationCredentials], Depends(bearer_scheme)],
) -> CurrentUserContext:
    token = _extract_authorization_bearer(request, credentials)
    current_user = _resolve_current_user(token, request)
    request.state.current_user = current_user
    request.state.user_jwt = token
    return current_user


CurrentUser = Annotated[CurrentUserContext, Depends(get_current_user)]


async def get_tenant(
    request: Request,
    current_user: CurrentUser,
) -> TenantContext:
    token = getattr(request.state, "user_jwt", "")
    if not token:
        raise HTTPException(status_code=500, detail="Internal server error")

    tenant = _build_tenant_context(current_user, token)
    request.state.tenant = tenant
    return tenant


Tenant = Annotated[TenantContext, Depends(get_tenant)]


def _normalize_secret(value: Optional[str]) -> str:
    return (value or "").strip().strip('"').strip("'").strip()


def parse_bearer_token(auth_header: Optional[str]) -> str:
    auth = (auth_header or "").strip()
    if not auth:
        raise HTTPException(status_code=401, detail="Unauthorized")

    normalized = auth
    lower = normalized.lower()

    if lower.startswith("bearerbearer"):
        normalized = f"Bearer {normalized[len('bearerbearer'):].lstrip(' :')}"
        lower = normalized.lower()
    elif lower.startswith("bearer:"):
        normalized = f"Bearer {normalized.split(':', 1)[1].strip()}"
        lower = normalized.lower()

    parts = normalized.split(None, 1)
    if len(parts) < 2:
        raise HTTPException(status_code=401, detail="Unauthorized")

    scheme = parts[0].rstrip(":")
    token = parts[1].strip()
    if scheme.lower() != "bearer":
        raise HTTPException(status_code=401, detail="Unauthorized")

    while token.lower().startswith("bearer "):
        token = token[7:].strip()

    token = token.strip().strip('"').strip("'").strip()
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")

    return token


def _extract_authorization_bearer(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = None,
) -> str:
    auth_header = request.headers.get("Authorization")
    if not auth_header and credentials and credentials.scheme and credentials.credentials:
        auth_header = f"{credentials.scheme} {credentials.credentials}"

    token = parse_bearer_token(auth_header)
    engine_secret = _normalize_secret(os.getenv("ENGINE_SECRET"))
    service_key = _normalize_secret(SUPABASE_SERVICE_ROLE_KEY)
    equals_engine = bool(engine_secret) and secrets.compare_digest(token, engine_secret)
    equals_service = bool(service_key) and secrets.compare_digest(token, service_key)

    if equals_service:
        _auth_log(logging.WARNING, "service_role_token_received", request=request)
    elif equals_engine:
        _auth_log(logging.INFO, "engine_secret_token_received", request=request)

    return token


async def get_import_tenant(
    request: Request,
    credentials: Annotated[Optional[HTTPAuthorizationCredentials], Depends(bearer_scheme)],
) -> TenantContext:
    token = _extract_authorization_bearer(request, credentials)
    engine_secret = _normalize_secret(os.getenv("ENGINE_SECRET"))

    if engine_secret and secrets.compare_digest(token, engine_secret):
        company_id = (request.headers.get("X-Company-Id") or "").strip()
        if not company_id:
            raise HTTPException(
                status_code=400,
                detail="X-Company-Id header is required for engine authentication.",
            )

        try:
            company_uuid = str(uuid.UUID(company_id))
        except ValueError:
            raise HTTPException(status_code=400, detail="X-Company-Id must be a valid UUID.")
        comp_res = (
            admin_client.table("companies")
            .select("id, installation_limit")
            .eq("id", company_uuid)
            .limit(1)
            .execute()
        )
        if not comp_res.data:
            raise HTTPException(status_code=403, detail="Forbidden")
        max_inst = comp_res.data[0].get("installation_limit") or 500

        tenant = TenantContext(
            user_id="engine_service",
            company_id=company_id,
            jwt="",
            installation_limit=max_inst,
        )
        request.state.tenant = tenant
        return tenant

    current_user = _resolve_current_user(token, request)
    request.state.current_user = current_user
    request.state.user_jwt = token
    tenant = _build_tenant_context(current_user, token)
    request.state.tenant = tenant
    return tenant


ImportTenant = Annotated[TenantContext, Depends(get_import_tenant)]


async def get_auth_context(
    request: Request,
    credentials: Annotated[Optional[HTTPAuthorizationCredentials], Depends(bearer_scheme)],
) -> TenantContext:
    return await get_import_tenant(request, credentials)


AuthTenant = Annotated[TenantContext, Depends(get_auth_context)]


async def get_internal_metrics_auth(
    request: Request,
    credentials: Annotated[Optional[HTTPAuthorizationCredentials], Depends(bearer_scheme)],
) -> TenantContext:
    token = _extract_authorization_bearer(request, credentials)
    engine_secret = _normalize_secret(os.getenv("ENGINE_SECRET"))

    if engine_secret and secrets.compare_digest(token, engine_secret):
        tenant = TenantContext(
            user_id="engine_service",
            company_id="internal_metrics",
            jwt="",
            installation_limit=0,
        )
        request.state.tenant = tenant
        return tenant

    current_user = _resolve_current_user(token, request)
    request.state.current_user = current_user
    request.state.user_jwt = token
    tenant = _build_tenant_context(current_user, token)
    request.state.tenant = tenant
    return tenant


InternalMetricsTenant = Annotated[TenantContext, Depends(get_internal_metrics_auth)]

class LocationTypeEnum(str, Enum):
    residential = "residential"
    industrial = "industrial"

class InstallationCreate(BaseModel):
    client_alias: Optional[str] = None
    installation_year: int
    kwp: float
    inverter_model: str = "Unknown"
    has_battery: bool = False
    location_type: LocationTypeEnum = LocationTypeEnum.residential
    tariff_type: str = "standard"
    estimated_consumption: float = 0
    dc_ac_ratio: float = 1.0
    has_maintenance_contract: bool = False
    country: str = "Unknown"

InstallationCreate.model_rebuild()

class OpportunityTypeEnum(str, Enum):
    battery_upgrade = "battery_upgrade"
    inverter_replacement = "inverter_replacement"
    system_expansion = "system_expansion"
    maintenance = "maintenance"
    ev_charger = "ev_charger"

class TrackingUpdate(BaseModel):
    installation_id: str
    opportunity_type: OpportunityTypeEnum
    contacted: bool = False
    result: str = ""
    closed: bool = False
    value: float = 0.0

class StatusUpdatePayload(BaseModel):
    status: Literal["New", "Contacted", "Proposal", "Closed", "Lost"]
    closed_value: float = 0.0

class NotesUpdatePayload(BaseModel):
    notes: str

class SalesActionTypeEnum(str, Enum):
    called = "called"
    email_sent = "email_sent"
    proposal_sent = "proposal_sent"

class SalesActionPayload(BaseModel):
    action: SalesActionTypeEnum
    note: str = ""

# Ensure models are fully resolved before FastAPI builds OpenAPI.
for _model in (
    TenantContext,
    InstallationCreate,
    TrackingUpdate,
    StatusUpdatePayload,
    NotesUpdatePayload,
    SalesActionPayload,
):
    _model.model_rebuild()

# ─── Helpers ──────────────────────────────────────────────────────────────────
def opportunity_display_es(opportunity_slug: Optional[str]) -> str:
    if not opportunity_slug:
        return "Oportunidad Comercial"
    if opportunity_slug in OPPORTUNITY_DISPLAY_ES:
        return OPPORTUNITY_DISPLAY_ES[opportunity_slug]
    return opportunity_slug.replace("_", " ").title()


def pipeline_status_display_es(status_value: Optional[str]) -> str:
    if not status_value:
        return "Sin estado"
    return PIPELINE_STATUS_ES.get(status_value, status_value)


def build_sales_email_draft(client: Dict) -> Dict[str, str]:
    alias = client.get("client_alias") or "Cliente"
    opportunity_slug = client.get("opportunity_type")
    opportunity_name = opportunity_display_es(opportunity_slug)
    expected_value = float(client.get("expected_value") or 0)
    estimated_savings = float(client.get("estimated_battery_savings") or 0)
    payback_years = client.get("battery_payback_years")
    sales_script_short = (client.get("sales_script_short") or "").strip()

    savings_line = ""
    if estimated_savings > 0:
        savings_line = f"Hemos detectado un ahorro potencial de aproximadamente {round(estimated_savings):,.0f} EUR al año."
    else:
        savings_line = "Hemos detectado una oportunidad de mejora con impacto económico positivo para su instalación."

    payback_line = ""
    if payback_years:
        payback_line = f"La amortización estimada se sitúa en torno a {payback_years} años."

    script_line = ""
    if sales_script_short:
        script_line = "Nuestro análisis indica que este sistema es un buen candidato para activar esta mejora en el corto plazo."

    subject = f"Propuesta de mejora energética para {alias}: {opportunity_name}"
    body = dedent(
        f"""\
        Hola,

        Soy del equipo técnico-comercial de su instalador solar.

        Durante una revisión reciente de su sistema, detectamos una oportunidad prioritaria: {opportunity_name}.
        {savings_line}
        {payback_line}
        {script_line}

        Valor potencial estimado de esta oportunidad: {round(expected_value):,.0f} EUR.

        Si le encaja, podemos enviarle una propuesta personalizada y resolver cualquier duda en una llamada de 10 minutos.

        Quedo atento para coordinarlo.
        """
    ).strip()

    return {
        "email_subject": subject,
        "email_body": body,
    }


def scoped_client(jwt: str) -> Client:
    """Creates a per-request client with the user's JWT (activates RLS auth.uid())."""
    client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    client.postgrest.auth(jwt)
    return client

# ─── App & Rate Limiter ────────────────────────────────────────────────────────
def _tenant_key(request: Request) -> str:
    tenant: Optional[TenantContext] = getattr(request.state, "tenant", None)
    return tenant.user_id if tenant else get_remote_address(request)


limiter = Limiter(key_func=_tenant_key, default_limits=["60/minute"])

# ------------------------------------
# SCORING ENGINE
# ------------------------------------

ENGINE_SECRET = _normalize_secret(os.getenv("ENGINE_SECRET"))
scoring_lock = Lock()

def core_score_all_installations():
    if not scoring_lock.acquire(blocking=False):
        logger.warning("ENGINE: Scoring is already running. Skipping overlapping run.")
        return {"message": "Scoring already running", "skipped": True}

    start_time = datetime.now(timezone.utc)
    logger.info("ENGINE: Starting full opportunity scoring run.")

    try:
        # Fresh admin client to avoid stale schema cache
        fresh_admin: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)

        # PREP FOR BLOCK 1: Set calculated_month to first day of current month
        calculated_month = start_time.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        now_year = start_time.year

        # 2. Prevent Full-Table Scan: Fetch companies first
        comp_res = fresh_admin.table("companies").select("id, name").execute()
        companies = comp_res.data or []
        
        total_companies = len(companies)
        total_installations_scored = 0
        companies_failed = 0
        
        logger.info(f"ENGINE: Found {total_companies} companies to process.")

        # 3. Iterate Per Company (Tenant Isolation)
        for company in companies:
            company_id = company["id"]
            company_name = company.get("name", "Unknown")
            
            try:
                logger.info(f"ENGINE: Processing company {company_name} ({company_id})")
                
                # BLOCK 2: Fetch Company Parameters (1 fetch per tenant)
                try:
                    comp_param_res = fresh_admin.table("company_parameters").select("*").eq("company_id", company_id).execute()
                    comp_params = comp_param_res.data[0] if comp_param_res.data else {}
                except Exception:
                    comp_params = {}
                    
                # Resolve weights, use defaults if missing
                battery_weight = float(comp_params.get("battery_weight", 1.0) if comp_params.get("battery_weight") is not None else 1.0)
                maintenance_weight = float(comp_params.get("maintenance_weight", 1.0) if comp_params.get("maintenance_weight") is not None else 1.0)
                expansion_weight = float(comp_params.get("expansion_weight", 1.0) if comp_params.get("expansion_weight") is not None else 1.0)
                ev_weight = float(comp_params.get("ev_weight", 1.0) if comp_params.get("ev_weight") is not None else 1.0)
                industrial_weight = float(comp_params.get("industrial_weight", 1.0) if comp_params.get("industrial_weight") is not None else 1.0)
                weights = {
                    "battery": battery_weight,
                    "maintenance": maintenance_weight,
                    "expansion": expansion_weight,
                    "ev": ev_weight,
                    "industrial": industrial_weight
                }

                offset = 0
                had_installations = False
                while True:
                    inst_res = (
                        fresh_admin.table("installations")
                        .select("*")
                        .eq("company_id", company_id)
                        .range(offset, offset + SCORING_BATCH_SIZE - 1)
                        .execute()
                    )
                    installations = inst_res.data or []

                    if not installations:
                        if not had_installations:
                            logger.info(f"ENGINE: No installations found for company {company_name}")
                        break

                    had_installations = True
                    upsert_payload = []
                    clients_payload = []
                    
                    for inst in installations:
                        result = compute_opportunity_score(
                            installation=inst,
                            weights=weights,
                            now_year=now_year,
                            calculated_month=calculated_month
                        )

                        # BLOQUE A.3: solo generar oportunidades comerciales con score >= 40
                        if not result.get("is_opportunity", False):
                            continue
                        
                        result_copy = result.copy()
                        result_copy.pop("is_opportunity", None)
                        annual_export = result_copy.pop("estimated_annual_export_kwh", 0)
                        battery_savings = result_copy.pop("estimated_battery_savings", 0)
                        payback = result_copy.pop("battery_payback_years", 0)
                        batt_score = result_copy.pop("battery_opportunity_score", 0)
                        sales_script_long = result_copy.pop("sales_script_long", "")
                        sales_script_short = result_copy.pop("sales_script_short", "")
                        opportunity_reason = result_copy.pop("opportunity_reason", "")

                        # Compatibilidad con tenants donde opportunity_scores aún no tiene columnas añadidas
                        score_row = result_copy.copy()
                        score_row.pop("close_probability", None)
                        score_row.pop("priority_score", None)
                        score_row.pop("recommendation_level", None)
                        score_row.pop("inverter_score", None)
                        score_row.pop("value_breakdown", None)  # Remove breakdown from score storage
                        upsert_payload.append(score_row)

                        # Use expected_value from scoring engine (dynamic calculation)
                        score = result["total_score"]
                        opp_type = result["primary_reason"]
                        expected_value = result.get("expected_value", 0.0)

                        # Close probability from scoring engine
                        close_prob = result.get("close_probability", 0.1)

                        # Dynamic generation of client_alias using UUID snippet
                        client_alias = f"PV-{str(inst.get('id', '0000'))[:8].upper()}"

                        weighted_expected_revenue = expected_value * close_prob

                        # Priority Score from scoring engine
                        priority_score = result.get("priority_score", score)

                        clients_payload.append({
                            "company_id": inst["company_id"],
                            "client_alias": client_alias,
                            "anonymous_client": True,
                            "system_size_kwp": inst.get("kwp"),
                            "installation_year": inst.get("installation_year"),
                            "location_type": inst.get("location_type"),
                            "opportunity_type": opp_type,
                            "score": score,
                            "priority_score": round(priority_score, 1),
                            "expected_value": round(expected_value, 2),
                            "close_probability": close_prob,
                            "weighted_expected_revenue": round(weighted_expected_revenue, 2),
                            # Battery fields
                            "estimated_annual_export_kwh": annual_export,
                            "estimated_battery_savings": round(float(battery_savings or 0)),
                            "battery_payback_years": round(float(payback or 0), 1),
                            "battery_opportunity_score": batt_score,
                            "sales_script_long": sales_script_long,
                            "sales_script_short": sales_script_short,
                            "opportunity_reason": opportunity_reason,
                        })
                    
                    if upsert_payload:
                        # 1. Upsert Opportunity Scores (Legacy Engine Logic)
                        fresh_admin.table("opportunity_scores").upsert(
                            upsert_payload, 
                            on_conflict="company_id,installation_id,calculated_month"
                        ).execute()
                        
                        # 2. Upsert Commercial Pipeline (BLOCK R2 pipeline logic)
                        clients_upsert_res = fresh_admin.table("clients").upsert(
                            clients_payload,
                            on_conflict="company_id,client_alias"
                        ).execute()
                        
                        # 3. Create Auto Alerts for High Value Opportunities
                        if clients_upsert_res.data:
                            thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
                            existing_alerts_res = fresh_admin.table("opportunity_alerts").select("client_id").eq("company_id", company_id).eq("alert_type", "battery_opportunity").gte("created_at", thirty_days_ago).execute()
                            existing_alert_client_ids = {row["client_id"] for row in (existing_alerts_res.data or [])}
                            
                            alerts_payload = []
                            for c_row in clients_upsert_res.data:
                                score = float(c_row.get("battery_opportunity_score") or 0)
                                if score > 80:
                                    client_id = c_row["id"]
                                    alias = c_row.get("client_alias", "Unknown")
                                    if client_id in existing_alert_client_ids:
                                        logger.info(f"ENGINE: Skipping duplicate alert for client {alias}")
                                        continue
                                        
                                    sav = c_row.get("estimated_battery_savings", 0)
                                    pay = c_row.get("battery_payback_years", 0)
                                    
                                    alerts_payload.append({
                                        "client_id": client_id,
                                        "company_id": c_row["company_id"],
                                        "alert_type": "battery_opportunity",
                                        "alert_message": f"High value battery opportunity detected for client {alias}.\nEstimated savings €{sav}.\nPayback {pay} years."
                                    })
                            if alerts_payload:
                                fresh_admin.table("opportunity_alerts").insert(alerts_payload).execute()
                        
                        total_installations_scored += len(upsert_payload)
                        logger.info(
                            "ENGINE: Scored %s installations for %s (batch_offset=%s).",
                            len(upsert_payload),
                            company_name,
                            offset,
                        )

                    if len(installations) < SCORING_BATCH_SIZE:
                        break
                    offset += SCORING_BATCH_SIZE
                    
            except Exception as e:
                import traceback
                logger.error(f"ENGINE: Failed to process company {company_name} ({company_id}): {str(e)}")
                logger.error(f"ENGINE TRACEBACK: {traceback.format_exc()}")
                companies_failed += 1
                # 4. Basic Transaction Safety Logic: Catch error and continue loop
        runtime_seconds = (datetime.now(timezone.utc) - start_time).total_seconds()
        logger.info(f"ENGINE: Run complete. Scored {total_installations_scored} installations across {total_companies - companies_failed} companies. Failed: {companies_failed}. Runtime: {runtime_seconds:.2f}s")
        
        try:
            fresh_admin.table("execution_tracking").insert({
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "runtime_seconds": round(runtime_seconds, 2),
                "installations_processed": total_installations_scored
            }).execute()
            logger.info("ENGINE: Logged execution to tracking table.")
        except Exception as e:
            logger.error(f"ENGINE: Failed to log execution to tracking: {e}")
        
        return {
            "message": "Monthly scoring completed", 
            "total_installations": total_installations_scored,
            "companies_processed": total_companies - companies_failed,
            "companies_failed": companies_failed,
            "runtime_seconds": round(runtime_seconds, 2)
        }
    finally:
        scoring_lock.release()

# ─── Lifespan: Startup / Shutdown ─────────────────────────────────────────────
from contextlib import asynccontextmanager

scheduler = BackgroundScheduler()

@asynccontextmanager
async def lifespan(app: FastAPI):
    # FIX #3: Initialize Supabase admin client at startup, not at module import time
    global admin_client
    admin_client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    logger.info("Supabase admin client initialized.")

    # FIX #2: Validate ENGINE_SECRET at startup (warn only — don't crash the server)
    if not os.getenv("ENGINE_SECRET"):
        logger.warning("ENGINE_SECRET not configured — /api/engine/score-all will return 403.")

    scheduler.add_job(core_score_all_installations, 'interval', hours=24)
    scheduler.start()
    logger.info("APScheduler initialized for daily scoring.")

    yield
    scheduler.shutdown()

app = FastAPI(title="Solvist Opportunity Intelligence", version="5.0.0", lifespan=lifespan)

# CORS must be attached to the production app instance immediately after app init.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.state.limiter = limiter


def _request_id_for(request: Request) -> str:
    existing = getattr(request.state, "request_id", "")
    if existing:
        return existing

    incoming = (request.headers.get(REQUEST_ID_HEADER) or "").strip()
    request_id = incoming[:128] or str(uuid.uuid4())
    request.state.request_id = request_id
    return request_id


def rate_limit_handler(request: Request, exc: RateLimitExceeded) -> JSONResponse:
    request_id = _request_id_for(request)
    return JSONResponse(
        status_code=429,
        content={"detail": "Rate limit exceeded", "request_id": request_id},
        headers={REQUEST_ID_HEADER: request_id},
    )


app.add_exception_handler(RateLimitExceeded, rate_limit_handler)
app.add_middleware(SlowAPIMiddleware)

@app.post("/api/engine/score-all")
@limiter.limit("5/minute")
def score_all_installations(request: Request):
    """
    Monthly Serverless Job endpoint. Can be triggered manually.
    """
    provided = _normalize_secret(request.headers.get("X-ENGINE-SECRET"))
    if not provided or not ENGINE_SECRET or not secrets.compare_digest(provided, ENGINE_SECRET):
        logger.warning(f"ENGINE: Authentication failed from IP {get_remote_address(request)}")
        raise HTTPException(status_code=403, detail="Forbidden")

    # Trigger scoring job immediately via APScheduler (non-blocking)
    scheduler.add_job(core_score_all_installations, "date")
    return {"message": "Scoring triggered"}

@app.get("/api/portfolio-opportunity-value")
@limiter.limit("20/minute")
def get_portfolio_opportunity_value(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    # 1. Fetch all clients for the company
    res = (
        db.table("clients")
        .select("opportunity_type, expected_value")
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not res.data:
        return {
            "battery_upgrades_value": 0,
            "inverter_replacements_value": 0,
            "system_expansions_value": 0,
            "total_opportunity_value": 0,
            "systems_analyzed": 0
        }
    
    # 2. Map and aggregate
    battery_val = 0
    inverter_val = 0
    expansion_val = 0
    total_val = 0
    
    for client in res.data:
        val = float(client.get("expected_value") or 0)
        opp_type = client.get("opportunity_type")
        
        if opp_type == OPP_BATTERY_UPGRADE:
            battery_val += val
        elif opp_type == OPP_INVERTER_REPLACEMENT:
            inverter_val += val
        elif opp_type == OPP_SYSTEM_EXPANSION:
            expansion_val += val
            
        total_val += val
        
    return {
        "battery_upgrades_value": battery_val,
        "inverter_replacements_value": inverter_val,
        "system_expansions_value": expansion_val,
        "total_opportunity_value": total_val,
        "systems_analyzed": len(res.data)
    }


@app.get("/api/recontact-opportunities")
@limiter.limit("20/minute")
def get_recontact_opportunities(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    current_year = datetime.now().year
    recontact_limit = current_year - 3
    
    # Rules: installation_year <= current_year - 3 AND score > 60
    res = (
        db.table("clients")
        .select("id, expected_value")
        .eq("company_id", tenant.company_id)
        .lte("installation_year", recontact_limit)
        .gt("score", 60)
        .execute()
    )
    
    if not res.data:
        return {
            "clients_to_recontact": 0,
            "estimated_revenue": 0
        }
        
    count = len(res.data)
    total_rev = sum(float(c.get("expected_value") or 0) for c in res.data)
    
    return {
        "clients_to_recontact": count,
        "estimated_revenue": total_rev
    }

@app.middleware("http")
async def audit_middleware(request: Request, call_next):
    start = datetime.now(timezone.utc)
    request_id = _request_id_for(request)
    origin = request.headers.get("origin", "")

    if request.url.path in {"/api/public/portfolio-scan", "/portfolio-scan"} or request.method == "OPTIONS":
        logger.info(
            "CORS inbound request_id=%s method=%s path=%s origin=%s",
            request_id,
            request.method,
            request.url.path,
            origin,
        )

    try:
        response = await call_next(request)
    except Exception:
        logger.exception(
            "Unhandled request error request_id=%s method=%s path=%s",
            request_id,
            request.method,
            request.url.path,
        )
        response = JSONResponse(
            status_code=500,
            content={"detail": "Internal server error", "request_id": request_id},
        )

    response.headers[REQUEST_ID_HEADER] = request_id
    tenant: Optional[TenantContext] = getattr(request.state, "tenant", None)
    duration_ms = int((datetime.now(timezone.utc) - start).total_seconds() * 1000)

    logger.info(
        "REQUEST request_id=%s method=%s path=%s status=%s duration_ms=%s",
        request_id,
        request.method,
        request.url.path,
        response.status_code,
        duration_ms,
    )

    if request.url.path in {"/api/public/portfolio-scan", "/portfolio-scan"} or request.method == "OPTIONS":
        logger.info(
            "CORS outbound request_id=%s status=%s allow_origin=%s",
            request_id,
            response.status_code,
            response.headers.get("access-control-allow-origin", ""),
        )

    if tenant and request.url.path.startswith("/api"):
        log_entry = {
            "id": str(uuid.uuid4()),
            "user_id": tenant.user_id,
            "company_id": tenant.company_id,
            "method": request.method,
            "endpoint": str(request.url.path),
            "status_code": response.status_code,
            "ip": request.headers.get("x-forwarded-for", get_remote_address(request)),
            "duration_ms": duration_ms,
            "created_at": start.isoformat(),
        }
        try:
            admin_client.table("audit_log").insert(log_entry).execute()
        except Exception:
            pass

    return response




# ─── Endpoints: Infrastructure ─────────────────────────────────────────────────


@app.post("/api/create-checkout-session")
@limiter.limit("5/minute")
def create_checkout_session(request: Request, tenant: Tenant):
    """
    Placeholder for Stripe Checkout.
    Required for Step 10: Returns a Stripe checkout URL.
    Actual implementation of pricing logic not required yet.
    """
    if not STRIPE_SECRET_KEY or STRIPE_SECRET_KEY == "sk_test_placeholder":
        logger.warning("Stripe secret key not configured or using placeholder.")
        return {"error": "Stripe configuration missing"}

    try:
        # In a real scenario, you'd create a session here.
        # This is a placeholder as requested.
        return {
            "status": "success",
            "checkout_url": "https://checkout.stripe.com/pay/placeholder_session_id"
        }
    except Exception as e:
        logger.error(f"Stripe session error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error creating checkout session")


# ─── Endpoints: Data Ingestion (Plan Enforcement) ──────────────────────────────
# Moved above


PUBLIC_SCAN_MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024
CSV_MAX_ROWS = 5000
CSV_PROCESSING_TIMEOUT_SECONDS = 5.0
CSV_INGESTION_METRICS: Dict[str, int] = {
    "csv_total_uploads": 0,
    "csv_malformed_detected": 0,
    "csv_valid_rows": 0,
    "csv_invalid_rows": 0,
}
CSV_INGESTION_METRICS_LOCK = Lock()

PORTFOLIO_SCAN_REQUIRED_COLUMNS_MESSAGE = (
    "We could not detect the required columns in your file.\n\n"
    "Required columns:\n"
    "- system size (kWp)\n"
    "- installation year\n\n"
    "Please upload a CSV, Excel file, or Google Sheet containing these fields."
)


REQUIRED_CANONICAL_FIELDS = ("system_size_kwp",)  # Only kwp is strictly required
DEFAULT_INSTALLATION_YEAR = None  # Will use current year - 5 if missing

# Canonical CSV schema aliases. First matching column is selected.
CANONICAL_COLUMN_ALIASES: Dict[str, List[str]] = {
    "system_size_kwp": [
        "system_size_kwp",
        "kwp",
        "size_kwp",
        "system_size",
        "installed_kwp",
        "power_kwp",
    ],
    "installation_year": [
        "installation_year",
        "year",
        "install_year",
        "installed_year",
        "commissioning_year",
        "commission_year",
        "installationdate",
    ],
    "country": ["country", "region", "location", "country_code"],
    "city": ["city", "ciudad", "municipality", "town"],
    "has_battery": ["has_battery", "battery", "battery_installed", "with_battery"],
    "inverter_model": ["inverter_model", "inverter_brand", "inverter", "inverter_model_name"],
    "location_type": ["location_type", "client_type", "type", "sector"],
    "client_name": ["client_name", "name", "customer", "client"],
    "client_alias": ["client_alias", "alias", "client_id"],
    "tariff_type": ["tariff_type", "tariff", "rate_type"],
    "estimated_consumption": ["estimated_consumption", "consumption_kwh", "annual_consumption", "consumption"],
    "dc_ac_ratio": ["dc_ac_ratio", "dcac_ratio", "dcac"],
    "has_maintenance_contract": ["has_maintenance_contract", "maintenance_contract", "with_maintenance"],
}


def _normalize_column_name(raw_name: object) -> str:
    name = str(raw_name or "").strip().lower()
    name = re.sub(r"[^\w]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def _normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed: Dict[str, str] = {}
    used_names: Dict[str, int] = {}
    for original in df.columns:
        normalized = _normalize_column_name(original)
        if not normalized:
            normalized = "column"
        suffix = used_names.get(normalized, 0)
        if suffix > 0:
            target = f"{normalized}_{suffix}"
        else:
            target = normalized
        used_names[normalized] = suffix + 1
        renamed[original] = target
    return df.rename(columns=renamed)


def _increment_csv_metric(metric_name: str, increment: int = 1) -> None:
    with CSV_INGESTION_METRICS_LOCK:
        current = int(CSV_INGESTION_METRICS.get(metric_name, 0))
        CSV_INGESTION_METRICS[metric_name] = current + increment


def _get_csv_ingestion_metrics_snapshot() -> Dict[str, int]:
    with CSV_INGESTION_METRICS_LOCK:
        return {
            "csv_total_uploads": int(CSV_INGESTION_METRICS.get("csv_total_uploads", 0)),
            "csv_malformed_detected": int(CSV_INGESTION_METRICS.get("csv_malformed_detected", 0)),
            "csv_valid_rows": int(CSV_INGESTION_METRICS.get("csv_valid_rows", 0)),
            "csv_invalid_rows": int(CSV_INGESTION_METRICS.get("csv_invalid_rows", 0)),
        }


def _ensure_csv_processing_time_budget(parse_started_at: float) -> None:
    elapsed = time.perf_counter() - parse_started_at
    if elapsed > CSV_PROCESSING_TIMEOUT_SECONDS:
        raise HTTPException(status_code=400, detail={"error": "CSV processing timeout"})


def _unwrap_line_wrapped_csv_text(text: str) -> str:
    """
    Some exports quote the entire row as a single CSV field:
    "col1,col2,col3"
    "v1,v2,v3"
    This unwraps only that malformed format while leaving valid CSV untouched.
    """
    lines = text.splitlines()
    if not lines:
        return text

    unwrapped_lines: List[str] = []
    wrapped_count = 0

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        if stripped.startswith('"') and stripped.endswith('"') and '","' not in stripped:
            candidate = stripped[1:-1]
            # Keep only plausible CSV-like rows.
            if any(delim in candidate for delim in [",", ";", "|", "\t"]):
                unwrapped_lines.append(candidate)
                wrapped_count += 1
                continue

        unwrapped_lines.append(stripped)

    if wrapped_count == 0:
        return text

    return "\n".join(unwrapped_lines)


def _attempt_csv_read(text: str, delimiter: str) -> pd.DataFrame:
    return pd.read_csv(
        io.StringIO(text),
        sep=delimiter,
        skipinitialspace=True,
        engine="python",
    )


def _looks_like_single_column_csv(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False
    if len(df.columns) != 1:
        return False

    only_col = str(df.columns[0])
    if any(delim in only_col for delim in [",", ";", "|", "\t"]):
        return True

    sample_values = df.iloc[:, 0].dropna().astype(str).head(10).tolist()
    return any(any(delim in value for delim in [",", ";", "|", "\t"]) for value in sample_values)


def _resolve_column_mapping(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    detected_columns = list(df.columns)
    logger.info("CSV columns detected: %s", detected_columns)

    for canonical, aliases in CANONICAL_COLUMN_ALIASES.items():
        source_column: Optional[str] = None
        for alias in aliases:
            normalized_alias = _normalize_column_name(alias)
            if normalized_alias in df.columns:
                source_column = normalized_alias
                break
        if source_column:
            mapping[canonical] = source_column
            logger.info("Mapped %s -> %s", canonical, source_column)

    for required_field in REQUIRED_CANONICAL_FIELDS:
        if required_field not in mapping:
            logger.warning("Missing %s", required_field)

    return mapping


def _coerce_required_float(
    raw_value: object,
    *,
    field_name: str,
    row_number: int,
) -> float:
    if raw_value is None or pd.isna(raw_value):
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": field_name, "row": row_number, "value": None},
        )

    normalized = str(raw_value).strip().lower()
    if not normalized:
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": field_name, "row": row_number, "value": str(raw_value)},
        )

    normalized = re.sub(r"[^0-9,.\-+]", "", normalized)
    if not normalized:
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": field_name, "row": row_number, "value": str(raw_value)},
        )

    if normalized.count(",") == 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", ".")
    elif normalized.count(",") > 0 and normalized.count(".") > 0:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif normalized.count(",") > 1 and normalized.count(".") == 0:
        normalized = normalized.replace(",", "")
    elif normalized.count(".") > 1 and normalized.count(",") == 0:
        normalized = normalized.replace(".", "")

    try:
        value = float(normalized)
    except Exception:
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": field_name, "row": row_number, "value": str(raw_value)},
        )

    return value


def _coerce_required_year(raw_value: object, *, row_number: int) -> int:
    value = _coerce_required_float(raw_value, field_name="installation_year", row_number=row_number)
    year = int(value)
    if not math.isclose(value, year, abs_tol=1e-6):
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": "installation_year", "row": row_number, "value": str(raw_value)},
        )
    current_year = datetime.now(timezone.utc).year + 1
    if year < 1900 or year > current_year:
        _increment_csv_metric("csv_invalid_rows")
        raise HTTPException(
            status_code=400,
            detail={"error": "Invalid required field value", "field": "installation_year", "row": row_number, "value": str(raw_value)},
        )
    return year


def _parse_bool_value(raw_value: object, default: bool = False) -> bool:
    if raw_value is None:
        return default
    if isinstance(raw_value, bool):
        return raw_value
    if pd.isna(raw_value):
        return default

    normalized = str(raw_value).strip().lower()
    if normalized in {"1", "true", "yes", "si", "sí"}:
        return True
    if normalized in {"0", "false", "no"}:
        return False
    return default


def _parse_float_value(raw_value: object, default: float = 0.0) -> float:
    if raw_value is None or pd.isna(raw_value):
        return default
    try:
        return float(raw_value)
    except Exception:
        try:
            normalized = str(raw_value).strip().replace(" ", "")
            if not normalized:
                return default

            if normalized.count(",") == 1 and normalized.count(".") == 0:
                normalized = normalized.replace(",", ".")
            elif normalized.count(",") > 0 and normalized.count(".") > 0:
                if normalized.rfind(",") > normalized.rfind("."):
                    normalized = normalized.replace(".", "").replace(",", ".")
                else:
                    normalized = normalized.replace(",", "")
            return float(normalized)
        except Exception:
            return default


def _normalize_location_type(raw_value: object) -> str:
    if raw_value is None or pd.isna(raw_value):
        return "residential"
    raw_loc = str(raw_value).lower().strip()
    if "commercial" in raw_loc or "industrial" in raw_loc:
        return "industrial"
    return "residential"


def _parse_installations_from_dataframe(
    df: pd.DataFrame,
    *,
    company_id: str,
    alias_prefix: str,
    required_columns_error_message: Optional[str] = None,
    parse_started_at: Optional[float] = None,
) -> List[Dict]:
    if parse_started_at is not None:
        _ensure_csv_processing_time_budget(parse_started_at)

    if df is None or df.empty:
        return []

    df = df.dropna(how="all")
    if df.empty:
        return []

    if len(df.index) > CSV_MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "CSV too large",
                "details": {
                    "max_rows": CSV_MAX_ROWS,
                },
            },
        )

    df = _normalize_dataframe_columns(df)
    mapping = _resolve_column_mapping(df)

    # Debug: log detected columns
    logger.info(f"CSV detected columns: {list(df.columns)}")
    logger.info(f"Column mapping: {mapping}")

    # Only system_size_kwp is strictly required
    if "system_size_kwp" not in mapping:
        if required_columns_error_message:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "Missing required column: kwp (system_size_kwp)",
                    "message": required_columns_error_message,
                    "columns_detected": list(df.columns),
                },
            )
        raise HTTPException(
            status_code=400,
            detail={
                "error": "Missing required column: kwp (system_size_kwp)",
                "columns_detected": list(df.columns),
            },
        )

    insert_payload: List[Dict] = []
    for row_pos, (_, row) in enumerate(df.iterrows()):
        if parse_started_at is not None and row_pos % 100 == 0:
            _ensure_csv_processing_time_budget(parse_started_at)

        row_number = row_pos + 2  # CSV/Excel with header row at line 1
        kwp = _coerce_required_float(
            row[mapping["system_size_kwp"]],
            field_name="system_size_kwp",
            row_number=row_number,
        )
        if kwp <= 0:
            _increment_csv_metric("csv_invalid_rows")
            raise HTTPException(
                status_code=400,
                detail={"error": "Invalid required field value", "field": "system_size_kwp", "row": row_number, "value": str(row[mapping["system_size_kwp"]])},
            )

        # installation_year is optional - default to None (will use current year - 5 in scoring)
        year_col = mapping.get("installation_year")
        if year_col and pd.notna(row.get(year_col)):
            try:
                year = int(str(row[year_col]).strip())
                if year < 1990 or year > 2100:
                    year = None
            except (ValueError, TypeError):
                year = None
        else:
            year = None

        alias_col = mapping.get("client_alias")
        location_col = mapping.get("location_type")
        battery_col = mapping.get("has_battery")
        tariff_col = mapping.get("tariff_type")
        consumption_col = mapping.get("estimated_consumption")
        dcac_col = mapping.get("dc_ac_ratio")
        maintenance_col = mapping.get("has_maintenance_contract")
        country_col = mapping.get("country")
        city_col = mapping.get("city")
        inverter_col = mapping.get("inverter_model")

        client_alias = (
            str(row[alias_col]).strip()
            if alias_col and pd.notna(row.get(alias_col)) and str(row[alias_col]).strip()
            else f"{alias_prefix}-{row_pos + 1:04d}"
        )
        stable_id_source = f"{company_id}:{client_alias}:{year or 'unknown'}:{kwp}"
        stable_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, stable_id_source))

        # Debug log for row data
        logger.debug(f"CSV row {row_number}: kwp={kwp}, year={year}, alias={client_alias}")

        # Get country with default
        country_value = "ES"  # Default to Spain
        if country_col and pd.notna(row.get(country_col)):
            country_value = str(row[country_col]).strip()

        insert_payload.append(
            {
                "id": stable_id,
                "system_size_kwp": float(kwp),
                "installation_year": int(year) if year else None,  # Will use default in scoring engine
                "country": country_value,
                "city": str(row[city_col]).strip() if city_col and pd.notna(row.get(city_col)) else None,
                "source": "csv_import",
                "raw_payload": {
                    "company_id": company_id,
                    "client_alias": client_alias,
                    "inverter_model": str(row[inverter_col]).strip() if inverter_col and pd.notna(row.get(inverter_col)) else None,
                    "location_type": _normalize_location_type(row[location_col]) if location_col else "residential",
                    "has_battery": _parse_bool_value(row[battery_col], False) if battery_col else False,
                    "tariff_type": str(row[tariff_col]).strip().lower() if tariff_col and pd.notna(row.get(tariff_col)) else "standard",
                    "estimated_consumption": _parse_float_value(row[consumption_col], 0.0) if consumption_col else 0.0,
                    "dc_ac_ratio": _parse_float_value(row[dcac_col], 1.0) if dcac_col else 1.0,
                    "has_maintenance_contract": _parse_bool_value(row[maintenance_col], False) if maintenance_col else False,
                },
            }
        )

    _increment_csv_metric("csv_valid_rows", len(insert_payload))
    return insert_payload


def _parse_installations_from_csv_bytes(
    csv_bytes: bytes,
    *,
    company_id: str,
    alias_prefix: str,
    required_columns_error_message: Optional[str] = None,
) -> List[Dict]:
    _increment_csv_metric("csv_total_uploads")
    if not csv_bytes:
        raise HTTPException(status_code=400, detail="CSV file is empty.")

    parse_started_at = time.perf_counter()
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    parse_error: Optional[Exception] = None
    df: Optional[pd.DataFrame] = None

    for encoding in encodings:
        _ensure_csv_processing_time_budget(parse_started_at)
        try:
            text = csv_bytes.decode(encoding)
        except UnicodeDecodeError as decode_error:
            parse_error = decode_error
            continue

        sample = text[:4096]
        delimiter = ","
        try:
            sniffed = csv.Sniffer().sniff(sample, delimiters=",;|\t")
            delimiter = sniffed.delimiter
        except Exception:
            if sample.count(";") > sample.count(","):
                delimiter = ";"

        try:
            df = _attempt_csv_read(text, delimiter)
            _ensure_csv_processing_time_budget(parse_started_at)

            # Fallback for malformed CSV exports where each full line is quoted.
            if _looks_like_single_column_csv(df):
                unwrapped_text = _unwrap_line_wrapped_csv_text(text)
                if unwrapped_text != text:
                    columns_before = len(df.columns)
                    sample_unwrapped = unwrapped_text[:4096]
                    unwrapped_delimiter = delimiter
                    try:
                        sniffed_unwrapped = csv.Sniffer().sniff(sample_unwrapped, delimiters=",;|\t")
                        unwrapped_delimiter = sniffed_unwrapped.delimiter
                    except Exception:
                        if sample_unwrapped.count(";") > sample_unwrapped.count(","):
                            unwrapped_delimiter = ";"
                    df = _attempt_csv_read(unwrapped_text, unwrapped_delimiter)
                    _increment_csv_metric("csv_malformed_detected")
                    logger.warning(
                        "CSV malformed detected -> fallback parser used columns_before=%s columns_after=%s rows=%s",
                        columns_before,
                        len(df.columns),
                        len(df.index),
                    )
                    _ensure_csv_processing_time_budget(parse_started_at)
            break
        except Exception as read_error:
            parse_error = read_error
            continue

    if df is None:
        raise HTTPException(
            status_code=400,
            detail="We could not parse your CSV file. Please verify encoding (UTF-8/Latin-1) and delimiter (comma or semicolon).",
        )

    return _parse_installations_from_dataframe(
        df,
        company_id=company_id,
        alias_prefix=alias_prefix,
        required_columns_error_message=required_columns_error_message,
        parse_started_at=parse_started_at,
    )


def _parse_installations_from_excel_bytes(
    excel_bytes: bytes,
    *,
    company_id: str,
    alias_prefix: str,
    required_columns_error_message: Optional[str] = None,
) -> List[Dict]:
    workbook = load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            return []

        headers: List[str] = []
        for idx, value in enumerate(header_row):
            if value is None:
                headers.append(f"column_{idx}")
                continue
            header = _normalize_column_name(value)
            headers.append(header or f"column_{idx}")

        records: List[Dict] = []
        for row in row_iter:
            if row is None:
                continue
            if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                continue

            record: Dict = {}
            for idx, header in enumerate(headers):
                record[header] = row[idx] if idx < len(row) else None
            records.append(record)
    finally:
        workbook.close()

    if not records:
        return []

    df = pd.DataFrame(records)
    return _parse_installations_from_dataframe(
        df,
        company_id=company_id,
        alias_prefix=alias_prefix,
        required_columns_error_message=required_columns_error_message,
    )


def _extract_google_sheet_parts(google_sheet_url: str) -> Dict[str, Optional[str]]:
    parsed = urlparse((google_sheet_url or "").strip())
    if parsed.scheme not in {"http", "https"}:
        raise HTTPException(status_code=400, detail="Invalid Google Sheets URL.")

    netloc = parsed.netloc.lower()
    if not netloc.endswith("docs.google.com"):
        raise HTTPException(status_code=400, detail="Google Sheets URL must be from docs.google.com.")

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", parsed.path)
    if not match:
        raise HTTPException(status_code=400, detail="Could not extract Google Sheets id from URL.")

    sheet_id = match.group(1)
    query = parse_qs(parsed.query)
    fragment_query = parse_qs(parsed.fragment)
    gid = query.get("gid", [None])[0] or fragment_query.get("gid", [None])[0]
    return {"sheet_id": sheet_id, "gid": gid}


def _download_google_sheet_csv_bytes(google_sheet_url: str, max_bytes: int) -> bytes:
    parts = _extract_google_sheet_parts(google_sheet_url)
    export_url = f"https://docs.google.com/spreadsheets/d/{parts['sheet_id']}/export?format=csv"
    if parts.get("gid"):
        export_url = f"{export_url}&gid={parts['gid']}"

    request = UrlRequest(export_url, headers={"User-Agent": "SolvistPortfolioScan/1.0"})
    try:
        with urlopen(request, timeout=15) as response:
            chunks: List[bytes] = []
            total = 0
            while True:
                chunk = response.read(64 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise HTTPException(status_code=413, detail="File too large. Maximum allowed size is 5MB.")
                chunks.append(chunk)

            content = b"".join(chunks)
            if not content:
                raise HTTPException(status_code=400, detail="Google Sheet returned empty data.")
            return content
    except HTTPError as e:
        raise HTTPException(status_code=400, detail=f"Unable to fetch Google Sheet (HTTP {e.code}).")
    except URLError as e:
        raise HTTPException(status_code=400, detail=f"Unable to fetch Google Sheet ({e.reason}).")


def _detect_currency_from_ip(ip_address: str) -> str:
    fallback_currency = "USD"
    if not ip_address:
        return fallback_currency

    lookup_url = f"https://ipapi.co/{ip_address}/json"
    request = UrlRequest(lookup_url, headers={"User-Agent": "SolvistPortfolioScan/1.0"})
    try:
        with urlopen(request, timeout=4) as response:
            raw_payload = response.read()
    except Exception as geolocation_error:
        logger.warning(f"Currency geolocation request failed: {geolocation_error}")
        return fallback_currency

    try:
        payload = json.loads(raw_payload.decode("utf-8"))
    except Exception as parsing_error:
        logger.warning(f"Currency geolocation payload parse failed: {parsing_error}")
        return fallback_currency

    if payload.get("error"):
        return fallback_currency

    currency = str(payload.get("currency") or "").strip().upper()
    if len(currency) == 3 and currency.isalpha():
        return currency

    country_code = str(payload.get("country_code") or "").strip().upper()
    return COUNTRY_TO_CURRENCY_FALLBACK.get(country_code, fallback_currency)


def _read_platform_scan_totals() -> Dict[str, float]:
    total_portfolios_analyzed = 0
    total_revenue_detected = 0.0
    page_size = 1000
    start = 0

    while True:
        end = start + page_size - 1
        res = (
            admin_client.table("portfolio_scans")
            .select("total_opportunity_value")
            .range(start, end)
            .execute()
        )
        rows = res.data or []
        if not rows:
            break

        total_portfolios_analyzed += len(rows)
        total_revenue_detected += sum(float(row.get("total_opportunity_value") or 0) for row in rows)

        if len(rows) < page_size:
            break
        start += page_size

    return {
        "total_portfolios_analyzed": total_portfolios_analyzed,
        "total_revenue_detected": round(total_revenue_detected, 2),
    }


@app.post("/api/installations")
@limiter.limit("20/minute")
def create_installation(
    request: Request,
    payload: Annotated[InstallationCreate, Body(...)],
    tenant: Tenant,
):
    if payload.installation_year < 2000:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Installation year must be >= 2000.")
        
    db = scoped_client(tenant.jwt)

    # Enforce plan limits
    count_res = db.table("installations").select("id", count="exact").eq("company_id", tenant.company_id).execute()
    current_count = count_res.count if hasattr(count_res, "count") and count_res.count is not None else len(count_res.data)

    if current_count >= tenant.installation_limit:
        raise HTTPException(status_code=402, detail=f"Plan limit exceeded. Max: {tenant.installation_limit}")

    # Pydantic implicitly formats Enums (payload.dict() calls their .value during JSON serialization via fastapi by default, or explicit mapping)
    data = payload.dict(exclude_none=True)
    data.pop("client_alias", None)
    data["location_type"] = payload.location_type.value
    data["company_id"] = tenant.company_id
    res = db.table("installations").insert(data).execute()
    return res.data[0]


# ─── Endpoints: Import Installations V1 ──────────────────────────────────────────
@app.post("/api/import/installations")
@limiter.limit("5/minute")
async def import_installations(request: Request, tenant: ImportTenant, file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files are supported.")
        
    db = admin_client if tenant.user_id == "engine_service" else scoped_client(tenant.jwt)
    logger.info("IMPORT: Received CSV upload filename=%s company_id=%s", file.filename, tenant.company_id)
    import_started_at = datetime.now(timezone.utc)
    
    try:
        content = await file.read()
        if len(content) > IMPORT_MAX_FILE_SIZE_BYTES:
            raise HTTPException(status_code=413, detail="File too large. Maximum allowed size is 10MB.")

        insert_payload = _parse_installations_from_csv_bytes(
            content,
            company_id=tenant.company_id,
            alias_prefix="PV",
        )
        logger.info("IMPORT: Parsed %s valid installations for company_id=%s", len(insert_payload), tenant.company_id)
        if not insert_payload:
            raise HTTPException(status_code=400, detail="No valid installations found in CSV")

        # Check plan limit against rows being imported (no company_id column on installations table)
        current_count = 0
        
        if current_count + len(insert_payload) > tenant.installation_limit:
            raise HTTPException(status_code=402, detail=f"Plan limit exceeded. Would reach {current_count + len(insert_payload)} but max is {tenant.installation_limit}")
            
        # Upsert in batches (dedupe by id)
        for chunk in [insert_payload[i:i + 100] for i in range(0, len(insert_payload), 100)]:
            db.table("installations").upsert(chunk, on_conflict="id").execute()
        logger.info("IMPORT: Inserted %s installations for company_id=%s", len(insert_payload), tenant.company_id)
            
        # Trigger engine job automatically after importing (non-blocking)
        scoring_triggered = False
        try:
            scheduler.add_job(core_score_all_installations, 'date')
            scoring_triggered = True
        except Exception as scoring_error:
            logger.warning("IMPORT: scoring trigger failed for company_id=%s error=%s", tenant.company_id, scoring_error)

        duration_seconds = (datetime.now(timezone.utc) - import_started_at).total_seconds()
        LAST_IMPORT_STATUS["last_import_rows"] = len(insert_payload)
        LAST_IMPORT_STATUS["last_import_duration"] = round(duration_seconds, 3)
        LAST_IMPORT_STATUS["last_import_opportunities"] = 0

        logger.info(
            "import_completed",
            extra={
                "company_id": tenant.company_id,
                "installations_imported": len(insert_payload),
                "opportunities_generated": 0,
                "duration_seconds": round(duration_seconds, 3),
                "scoring_triggered": scoring_triggered,
            },
        )
        
        return {
            "installations_imported": len(insert_payload),
            "scoring_triggered": scoring_triggered
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("IMPORT FAILED: %s", str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/portfolio-scan")
@app.post("/api/public/portfolio-scan")
@limiter.limit("10/minute")
async def public_portfolio_scan(
    request: Request,
    file: Optional[UploadFile] = File(None),
    google_sheet_url: Optional[str] = Form(None),
):
    filename = (file.filename or "").lower() if file else ""
    google_sheet_url = (google_sheet_url or "").strip()
    logger.info(
        "Public portfolio scan hit origin=%s ip=%s",
        request.headers.get("origin", ""),
        request.client.host if request.client else "unknown",
    )

    if not file and not google_sheet_url:
        raise HTTPException(
            status_code=400,
            detail="Provide a CSV/XLSX file or a public Google Sheets URL.",
        )

    if file and not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported.")

    forwarded_for = request.headers.get("x-forwarded-for")
    ip_address = forwarded_for.split(",")[0].strip() if forwarded_for else get_remote_address(request)

    # Cooldown protection by IP: one scan every 5 minutes
    try:
        last_scan_res = (
            admin_client.table("portfolio_scans")
            .select("created_at")
            .eq("ip_address", ip_address)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )

        last_scan_rows = last_scan_res.data or []
        if last_scan_rows:
            created_at_raw = last_scan_rows[0].get("created_at")
            if created_at_raw:
                created_at_str = str(created_at_raw)
                if created_at_str.endswith("Z"):
                    created_at_str = created_at_str[:-1] + "+00:00"

                last_scan_at = datetime.fromisoformat(created_at_str)
                if last_scan_at.tzinfo is None:
                    last_scan_at = last_scan_at.replace(tzinfo=timezone.utc)

                if datetime.now(timezone.utc) - last_scan_at < timedelta(minutes=5):
                    raise HTTPException(
                        status_code=429,
                        detail="Please wait a few minutes before running another scan.",
                    )
    except HTTPException:
        raise
    except Exception as cooldown_error:
        logger.warning(f"Public portfolio scan cooldown check failed: {cooldown_error}")

    content = b""
    try:
        if google_sheet_url:
            content = _download_google_sheet_csv_bytes(
                google_sheet_url,
                max_bytes=PUBLIC_SCAN_MAX_FILE_SIZE_BYTES,
            )
            installations = _parse_installations_from_csv_bytes(
                content,
                company_id="public-portfolio-scan",
                alias_prefix="SCAN",
                required_columns_error_message=PORTFOLIO_SCAN_REQUIRED_COLUMNS_MESSAGE,
            )
        else:
            content = await file.read() if file else b""
            if len(content) > PUBLIC_SCAN_MAX_FILE_SIZE_BYTES:
                raise HTTPException(
                    status_code=413,
                    detail="File too large. Maximum allowed size is 5MB.",
                )

            if filename.endswith(".xlsx"):
                installations = _parse_installations_from_excel_bytes(
                    content,
                    company_id="public-portfolio-scan",
                    alias_prefix="SCAN",
                    required_columns_error_message=PORTFOLIO_SCAN_REQUIRED_COLUMNS_MESSAGE,
                )
            else:
                installations = _parse_installations_from_csv_bytes(
                    content,
                    company_id="public-portfolio-scan",
                    alias_prefix="SCAN",
                    required_columns_error_message=PORTFOLIO_SCAN_REQUIRED_COLUMNS_MESSAGE,
                )

        if not installations:
            raise HTTPException(status_code=400, detail="No valid installations found in file")

        calculated_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        now_year = datetime.now(timezone.utc).year
        weights = {
            "battery": 1.0,
            "maintenance": 1.0,
            "expansion": 1.0,
            "ev": 1.0,
            "industrial": 1.0,
        }

        scored_clients: List[Dict] = []
        for inst in installations:
            result = compute_opportunity_score(
                installation=inst,
                weights=weights,
                now_year=now_year,
                calculated_month=calculated_month,
            )

            if not result.get("is_opportunity", False):
                continue

            opp_type = result.get("primary_reason")
            # Use dynamic expected_value from scoring engine
            expected_value = float(result.get("expected_value") or 0)

            scored_clients.append(
                {
                    "client_alias": inst.get("client_alias"),
                    "opportunity_type": opp_type,
                    "opportunity_type_display": opportunity_display_es(opp_type),
                    "score": float(result.get("total_score") or 0),
                    "priority_score": float(result.get("priority_score") or 0),
                    "close_probability": float(result.get("close_probability") or 0),
                    "expected_value": expected_value,
                }
            )

        scored_clients.sort(key=lambda row: row.get("priority_score", 0), reverse=True)
        weekly_priority_pool = scored_clients[:10]
        weekly_priority_clients = weekly_priority_pool[:3]

        systems_analyzed = len(installations)
        opportunities_detected = len(scored_clients)
        opportunity_rate = (
            (opportunities_detected / systems_analyzed) * 100
            if systems_analyzed > 0
            else 0.0
        )
        opportunity_rate = round(opportunity_rate, 1)
        total_opportunity_value = sum(float(row.get("expected_value") or 0) for row in scored_clients)
        average_opportunity_value = (
            total_opportunity_value / opportunities_detected
            if opportunities_detected > 0
            else 0.0
        )
        weekly_priority_value = sum(float(row.get("expected_value") or 0) for row in weekly_priority_pool)
        detected_currency = _detect_currency_from_ip(ip_address)

        try:
            admin_client.table("portfolio_scans").insert(
                {
                    "ip_address": ip_address,
                    "systems_analyzed": systems_analyzed,
                    "opportunities_detected": opportunities_detected,
                    "total_opportunity_value": round(total_opportunity_value, 2),
                }
            ).execute()
        except Exception as tracking_error:
            logger.warning(f"Public portfolio scan tracking failed: {tracking_error}")

        return {
            "systems_analyzed": systems_analyzed,
            "opportunities_detected": opportunities_detected,
            "opportunity_rate": opportunity_rate,
            "total_opportunity_value": round(total_opportunity_value, 2),
            "average_opportunity_value": round(average_opportunity_value, 2),
            "weekly_priority_clients": weekly_priority_clients,
            "weekly_priority_value": round(weekly_priority_value, 2),
            "currency": detected_currency,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Public portfolio scan error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error processing portfolio scan.")
    finally:
        if file:
            await file.close()
        content = b""


@app.get("/api/public/platform-stats")
@limiter.limit("30/minute")
def public_platform_stats(request: Request):
    try:
        return _read_platform_scan_totals()
    except Exception as e:
        logger.warning(f"Public platform stats failed: {e}")
        return {
            "total_portfolios_analyzed": 0,
            "total_revenue_detected": 0.0,
        }

# ─── Endpoints: Commercial Dashboard V1 ────────────────────────────────────────
@app.get("/api/dashboard")
@limiter.limit("30/minute")
def dashboard(request: Request, tenant: AuthTenant):
    db = admin_client if tenant.user_id == "engine_service" else scoped_client(tenant.jwt)

    # BLOCK 2: Fetch Active Threshold Parameter
    # Service role needed to read company_parameters as it has no RLS (engine/admin use)
    # or fallback to 65
    try:
        param_res = db.table("company_parameters").select("active_threshold").eq("company_id", tenant.company_id).execute()
        active_threshold = param_res.data[0].get("active_threshold") if param_res.data else 65
    except Exception:
        active_threshold = 65

    if active_threshold is None:
        active_threshold = 65

    current_month = date.today().replace(day=1).isoformat()

    # 1. Total installations (Postgres Count) - Unfiltered by month as it represents physical assets
    inst_count_res = db.table("installations").select("id", count="exact").eq("company_id", tenant.company_id).execute()
    total_inst = inst_count_res.count if hasattr(inst_count_res, "count") and inst_count_res.count is not None else len(inst_count_res.data)

    # 2. Total Scored Count (Isolated to current_month)
    total_scored_res = db.table("opportunity_scores").select("id", count="exact").eq("company_id", tenant.company_id).eq("calculated_month", current_month).execute()
    total_scored = total_scored_res.count if hasattr(total_scored_res, "count") and total_scored_res.count is not None else len(total_scored_res.data)

    # 3. Active Window Count (Isolated to current_month)
    active_res = (
        db.table("opportunity_scores")
        .select("id", count="exact")
        .eq("company_id", tenant.company_id)
        .eq("calculated_month", current_month)
        .gt("total_score", active_threshold)
        .execute()
    )
    active_window = active_res.count if hasattr(active_res, "count") and active_res.count is not None else len(active_res.data)

    window_pct = round((active_window / total_scored * 100), 1) if total_scored > 0 else 0

    # 4. Estimated potential value (Filtered query join on kwp, isolated to current_month)
    active_kwp_res = (
        db.table("opportunity_scores")
        .select("installations!inner(kwp)")
        .eq("company_id", tenant.company_id)
        .eq("calculated_month", current_month)
        .gt("total_score", active_threshold)
        .execute()
    )
    active_kwp = 0
    for row in active_kwp_res.data:
        inst = row.get("installations", {})
        if inst and isinstance(inst, dict):
            active_kwp += inst.get("kwp") or 0
    pot_value = round(active_kwp * 1500, 2)  # Proxy calculation

    # Top 5 clients (Isolated to current month)
    # Using Supabase joined query (installations is a FK in opportunity_scores)
    top_res = (
        db.table("opportunity_scores")
        .select("installation_id, total_score, primary_reason, recommended_action")
        .eq("company_id", tenant.company_id)
        .eq("calculated_month", current_month)
        .order("total_score", desc=True)
        .limit(5)
        .execute()
    )
    
    formatted_top = []
    for row in top_res.data:
        client_name = f"PV-{str(row.get('installation_id', ''))[:8].upper()}" if row.get("installation_id") else "Cliente"
        formatted_top.append({
            "client_name": client_name,
            "total_score": row["total_score"],
            "primary_reason": row["primary_reason"],
            "recommended_action": row["recommended_action"]
        })

    # Real Monthly Trend via Postgres RPC
    try:
        trend_res = db.rpc("get_monthly_avg_scores", {"p_company_id": tenant.company_id}).execute()
        monthly_trend = [float(row["avg_score"]) for row in trend_res.data] if trend_res.data else []
    except Exception as e:
        logger.error(f"Failed to fetch trend: {str(e)}")
        monthly_trend = []

    return {
        "total_installations": total_inst,
        "window_active_pct": window_pct,
        "estimated_potential_value": pot_value,
        "monthly_avg_score_trend": monthly_trend,
        "top_5_clients": formatted_top
    }


# ─── Endpoints: Activation List V1 ─────────────────────────────────────────────
@app.get("/api/activation")
@limiter.limit("30/minute")
def activation_list(request: Request, tenant: Tenant, limit: int = 20, offset: int = 0):
    db = scoped_client(tenant.jwt)
    current_month = date.today().replace(day=1).isoformat()
    res = (
        db.table("opportunity_scores")
        .select("id, installation_id, total_score, primary_reason, recommended_action, installations(location_type, installation_year)")
        .eq("company_id", tenant.company_id)
        .eq("calculated_month", current_month)
        .order("total_score", desc=True)
        .limit(limit)
        .offset(offset)
        .execute()
    )
    
    # Flatten the result mapping
    output = []
    
    # BLOCK F20: Fetch battery metrics from clients table
    client_aliases = [f"PV-{str(row['installation_id'])[:8].upper()}" for row in res.data if row.get('installation_id')]
    metrics_by_alias = {}
    if client_aliases:
        clients_res = db.table("clients").select("client_alias, estimated_annual_export_kwh, estimated_battery_savings, battery_payback_years, battery_opportunity_score").in_("client_alias", client_aliases).eq("company_id", tenant.company_id).execute()
        metrics_by_alias = {c["client_alias"]: c for c in clients_res.data}

    for row in res.data:
        inst = row.get("installations", {}) or {}
        alias = f"PV-{str(row['installation_id'])[:8].upper()}" if row.get("installation_id") else ""
        metrics = metrics_by_alias.get(alias, {})
        
        output.append({
            "score_id": row.get("id"),
            "installation_id": row.get("installation_id"),
            "client_name": alias or "Cliente",
            "location_type": inst.get("location_type"),
            "installation_year": inst.get("installation_year"),
            "total_score": row.get("total_score"),
            "primary_reason": row.get("primary_reason"),
            "recommended_action": row.get("recommended_action"),
            "estimated_annual_export_kwh": metrics.get("estimated_annual_export_kwh"),
            "estimated_battery_savings": metrics.get("estimated_battery_savings"),
            "battery_payback_years": metrics.get("battery_payback_years"),
            "battery_opportunity_score": metrics.get("battery_opportunity_score")
        })
    return output


# ─── Endpoints: Insights Panel V1 ──────────────────────────────────────────────
@app.get("/api/insights")
@limiter.limit("30/minute")
def insights(request: Request, tenant: AuthTenant):
    db = admin_client if tenant.user_id == "engine_service" else scoped_client(tenant.jwt)
    current_month = date.today().replace(day=1).isoformat()
    res = db.table("opportunity_scores").select("battery_score, maintenance_score, ev_score").eq("company_id", tenant.company_id).eq("calculated_month", current_month).execute()
    
    total = len(res.data)
    batt_opps = sum(1 for r in res.data if r.get("battery_score", 0) >= 15)
    maint_opps = sum(1 for r in res.data if r.get("maintenance_score", 0) >= 15)
    
    perc_batt = round((batt_opps / total * 100)) if total > 0 else 0
    perc_maint = round((maint_opps / total * 100)) if total > 0 else 0
    
    return [
        {"metric": "Oportunidades Batería", "value": f"{perc_batt}%", "description": "en ventana óptima (>15 pts)"},
        {"metric": "Riesgo Mantenimiento", "value": f"{perc_maint}%", "description": "instalaciones desprotegidas"},
    ]


@app.post("/api/tracking")
@limiter.limit("30/minute")
def add_tracking(request: Request, payload: TrackingUpdate, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    # Ownership check via RLS happens automatically during insert, but we verify existence
    check = db.table("installations").select("id").eq("id", payload.installation_id).single().execute()
    if not check.data:
        raise HTTPException(status_code=404, detail="Installation not found")
        
    data = payload.dict()
    data["company_id"] = tenant.company_id
    if payload.contacted:
        data["contact_date"] = datetime.now(timezone.utc).isoformat()
    if payload.closed:
        data["closed_at"] = datetime.now(timezone.utc).isoformat()
        
    res = db.table("execution_tracking").insert(data).execute()
    return {"message": "Tracking saved", "id": res.data[0]["id"]}


# ─── Endpoints: Minimalist PDF Generation V1 ───────────────────────────────────
@app.get("/api/activation/{installation_id}/pdf")
@limiter.limit("10/minute")
def generate_activation_pdf(request: Request, installation_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    current_month = date.today().replace(day=1).isoformat()
    
    # Verify access via RLS and join
    res = (
        db.table("opportunity_scores")
        .select("total_score, primary_reason, recommended_action, installations(*)")
        .eq("company_id", tenant.company_id)
        .eq("installation_id", installation_id)
        .eq("calculated_month", current_month)
        .single()
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Score/Installation not found or access denied")
        
    data = res.data
    inst = data.get("installations", {})
    
    # ── FPDF Logic ── 
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    pdf.set_fill_color(10, 14, 40)
    pdf.rect(0, 0, 210, 297, "F")
    
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(100, 180, 255)
    pdf.cell(0, 15, "SOLVIST - Oportunidad Comercial", ln=True, align="C")
    
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(220, 220, 220)
    activation_alias = f"PV-{str(installation_id)[:8].upper()}"
    pdf.cell(0, 10, f"Cliente: {activation_alias}", ln=True, align="C")
    pdf.ln(10)
    
    # Data card
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(100, 180, 255)
    pdf.cell(0, 10, "1. Datos del Sistema Analizado", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(180, 190, 210)
    pdf.cell(0, 8, f"Potencia Instalada: {inst.get('kwp', 0)} kWp", ln=True)
    pdf.cell(0, 8, f"Despliegue inicial: Año {inst.get('installation_year')}", ln=True)
    pdf.cell(0, 8, f"Tipo de consumo: {str(inst.get('location_type')).capitalize()}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(255, 100, 100) # Highlight
    pdf.cell(0, 10, "2. Oportunidad Detectada", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, f"Score Total: {data.get('total_score')} / 100 pts", ln=True)
    pdf.cell(0, 8, f"Vector principal: {data.get('primary_reason')}", ln=True)
    pdf.ln(5)
    
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(100, 255, 150) # Green tip
    pdf.cell(0, 10, "3. Acción Recomendada (Playbook)", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(220, 220, 220)
    pdf.multi_cell(0, 8, data.get('recommended_action'))
    
    pdf.ln(15)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, f"Documento generado autom\u00e1ticamente por el motor Solvist. (ID: {installation_id})", align="C")
    
    pdf_bytes = bytes(pdf.output())
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=oportunidad_{activation_alias}.pdf"})


@app.get("/api/client/{client_id}/proposal-pdf")
@limiter.limit("10/minute")
def generate_proposal_pdf(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    res = (
        db.table("clients")
        .select("*")
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .single()
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found or access denied")
        
    client = res.data
    
    # Generate Commercial Proposal PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # ─── Page Background ───────────────────────────────────────────
    pdf.set_fill_color(10, 14, 40)
    pdf.rect(0, 0, 210, 297, "F")
    
    # ─── Title ─────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 26)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 20, "Battery Upgrade Opportunity", ln=True, align="C")
    
    # ─── Client Info ───────────────────────────────────────────────
    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(180, 190, 255)
    pdf.cell(0, 10, f"Client: {client.get('client_alias')}", ln=True, align="C")
    pdf.ln(10)
    
    # ─── 1. Opportunity Detected ───────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(255, 100, 100)
    pdf.cell(0, 10, "1. Opportunity Detected", ln=True)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, f"Opportunity Type: {client.get('opportunity_type')}", ln=True)
    pdf.cell(0, 8, f"Expected Value: EUR {client.get('expected_value', 0):,.2f}", ln=True)
    pdf.cell(0, 8, f"Close Probability: {round(client.get('close_probability', 0) * 100)}%", ln=True)
    pdf.cell(0, 8, f"AI Score: {client.get('score')} pts", ln=True)
    pdf.ln(8)
    
    # ─── 2. Battery Analysis ───────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(100, 255, 150)
    pdf.cell(0, 10, "2. Battery Analysis", ln=True)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(220, 220, 220)
    
    export_kwh = client.get("estimated_annual_export_kwh", 0) or 0
    savings = client.get("estimated_battery_savings", 0) or 0
    payback = client.get("battery_payback_years", 0) or 0
    batt_score = client.get("battery_opportunity_score", 0) or 0
    
    pdf.cell(0, 8, f"Estimated Annual Export: {round(export_kwh):,} kWh/year", ln=True)
    pdf.cell(0, 8, f"Potential Battery Savings: EUR {round(savings):,}/year", ln=True)
    pdf.cell(0, 8, f"Estimated Payback Period: {payback} years", ln=True)
    pdf.cell(0, 8, f"Battery Opportunity Score: {batt_score}/100", ln=True)
    pdf.ln(8)
    
    # ─── 3. System Information ─────────────────────────────────────
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(100, 180, 255)
    pdf.cell(0, 10, "3. System Information", ln=True)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(220, 220, 220)
    pdf.cell(0, 8, f"System Size: {client.get('system_size_kwp') or 'N/A'} kWp", ln=True)
    pdf.cell(0, 8, f"Installation Year: {client.get('installation_year') or 'N/A'}", ln=True)
    pdf.cell(0, 8, f"Location Type: {str(client.get('location_type') or 'N/A').capitalize()}", ln=True)
    pdf.ln(8)
    
    # ─── 4. Opportunity Explanation ────────────────────────────────
    sales_script = client.get("sales_script_long") or ""
    if sales_script:
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(255, 200, 100)
        pdf.cell(0, 10, "4. Opportunity Explanation", ln=True)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(220, 220, 220)
        # Replace newlines and euro signs for fpdf compatibility
        clean_script = sales_script.replace("\n\n", "\n").replace("\u20ac", "EUR ").strip()
        pdf.multi_cell(0, 7, clean_script)
    
    # ─── Footer ────────────────────────────────────────────────────
    pdf.ln(20)
    pdf.set_font("Helvetica", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, f"Proposal generated by Solvist Intelligence Engine. (ID: {client_id})", align="C")
    
    pdf_bytes = bytes(pdf.output())
    
    # ─── Record Event ──────────────────────────────────────────────
    db.table("opportunity_events").insert({
        "client_id": client_id,
        "company_id": tenant.company_id,
        "event_type": "proposal_generated",
        "event_description": "Generated Commercial Proposal PDF"
    }).execute()
    
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=proposal_{client.get('client_alias')}.pdf"})


@app.get("/api/proposal/{client_id}/pdf")
@limiter.limit("10/minute")
def generate_proposal_pdf_alias(request: Request, client_id: str, tenant: Tenant):
    """Alias endpoint for proposal PDF generation."""
    return generate_proposal_pdf(request, client_id, tenant)

# ─── Endpoints: Commercial Revenue Intelligence V1 (BLOCK R3) ─────────────────

@app.get("/api/alerts")
@limiter.limit("60/minute")
def get_opportunity_alerts(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("opportunity_alerts")
        .select("*")
        .eq("company_id", tenant.company_id)
        .eq("seen", False)
        .order("created_at", desc=True)
        .execute()
    )
    return res.data or []

@app.post("/api/alerts/{alert_id}/seen")
@limiter.limit("60/minute")
def mark_alert_seen(alert_id: str, request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("opportunity_alerts")
        .update({"seen": True})
        .eq("id", alert_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    if not res.data:
        raise HTTPException(status_code=404, detail="Alert not found")
    return {"success": True, "alert": res.data[0]}

@app.get("/api/opportunity-insights")
@limiter.limit("30/minute")
def opportunity_insights(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("opportunity_type, expected_value, status")
        .eq("company_id", tenant.company_id)
        .neq("status", "Lost")
        .execute()
    )
    
    type_values = {}
    total_opps = 0
    
    for row in res.data:
        val = float(row.get("expected_value") or 0)
        slug = row.get("opportunity_type") or "unknown"
        type_values[slug] = type_values.get(slug, 0.0) + val
        total_opps += 1
        
    opportunities_list = [
        {
            "slug": k, 
            "type": OPP_DISPLAY_NAMES.get(k, k.replace("_", " ").title()), 
            "value": v
        } for k, v in type_values.items()
    ]
    opportunities_list.sort(key=lambda x: x["value"], reverse=True)
    
    return {
        "opportunities": opportunities_list,
        "total_opportunities": total_opps
    }

@app.get("/api/commercial-dashboard")
@limiter.limit("30/minute")
def commercial_dashboard(request: Request, tenant: Tenant):
    try:
        db = scoped_client(tenant.jwt)
        if not db:
            raise RuntimeError("Base de datos no inicializada")

        res = db.rpc("get_commercial_dashboard_metrics", {"p_company_id": tenant.company_id}).execute()

        opportunity_value_res = (
            db.table("clients")
            .select("expected_value")
            .eq("company_id", tenant.company_id)
            .gte("score", 40)
            .execute()
        )
        total_opportunity_value = sum(float(row.get("expected_value") or 0) for row in (opportunity_value_res.data or []))
        
        if res.data:
            data = res.data[0]
            return {
                "currency": data.get("currency", "EUR"),
                "total_systems": data.get("total_systems", 0),
                "total_pipeline_value": float(total_opportunity_value),
                "total_opportunity_value": float(total_opportunity_value),
                "weighted_forecast": float(data.get("weighted_forecast") or 0),
                "closed_revenue": float(data.get("closed_revenue") or 0),
                "hot_leads_count": data.get("hot_leads_count", 0)
            }
    except Exception as e:
        logger.error(f"Dashboard RPC Error: {str(e)}")
        # Devolver fallback con ceros para que el frontend no rompa si hay un error temporal
        
    return {
        "currency": "EUR",
        "total_systems": 0,
        "total_pipeline_value": 0.0,
        "total_opportunity_value": 0.0,
        "weighted_forecast": 0.0,
        "closed_revenue": 0.0,
        "hot_leads_count": 0
    }


@app.get("/api/top-priority")
@limiter.limit("30/minute")
def top_priority(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, close_probability, score, priority_score, status, battery_opportunity_score")
        .eq("company_id", tenant.company_id)
        .neq("status", "Closed")
        .execute()
    )
    # Sort by priority_score descending
    data = sorted(
        res.data,
        key=lambda c: (c.get("priority_score", 0) or 0),
        reverse=True
    )
    
    # Enrichment with display names
    for c in data:
        slug = c.get("opportunity_type")
        c["opportunity_type_display"] = opportunity_display_es(slug)
        
    return data[:5]


@app.get("/api/weekly-priority")
@limiter.limit("30/minute")
def weekly_priority(request: Request, tenant: Tenant, limit: int = 10, min_priority: float = 0):
    db = scoped_client(tenant.jwt)
    safe_limit = max(1, min(limit, 100))

    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, close_probability, priority_score, status")
        .eq("company_id", tenant.company_id)
        .gte("score", 40)
        .neq("status", "Closed")
        .gte("priority_score", min_priority)
        .order("priority_score", desc=True)
        .limit(safe_limit)
        .execute()
    )

    data = res.data or []
    for row in data:
        slug = row.get("opportunity_type")
        row["opportunity_type_display"] = opportunity_display_es(slug)
        row["status_display"] = pipeline_status_display_es(row.get("status"))

    return data


@app.get("/api/hot-leads")
@limiter.limit("30/minute")
def hot_leads(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("client_alias, opportunity_type, expected_value, close_probability, status")
        .eq("company_id", tenant.company_id)
        .gte("close_probability", 0.6)
        .neq("status", "Closed")
        .execute()
    )
    return res.data


@app.get("/api/revenue-at-risk")
@limiter.limit("30/minute")
def revenue_at_risk(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    threshold_date = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    
    res = (
        db.table("clients")
        .select("client_alias, opportunity_type, expected_value, status, last_contact_at")
        .eq("company_id", tenant.company_id)
        .neq("status", "Closed")
        .or_(f"last_contact_at.is.null,last_contact_at.lt.{threshold_date}")
        .execute()
    )
    return res.data


# Moved above

@app.post("/api/client/{client_id}/status")
@limiter.limit("30/minute")
def update_client_status(request: Request, client_id: str, payload: StatusUpdatePayload, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    update_data = {
        "status": payload.status,
        "status_updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if payload.status == "Closed":
        update_data["closed_value"] = payload.closed_value
    else:
        update_data["closed_value"] = 0.0
        
    res = (
        db.table("clients")
        .update(update_data)
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found or access denied")
        
    db.table("opportunity_events").insert({
        "client_id": client_id,
        "company_id": tenant.company_id,
        "event_type": "status_updated",
        "event_description": f"Status updated to {payload.status}"
    }).execute()
        
    return res.data[0]

@app.get("/api/client/{client_id}/timeline")
@limiter.limit("30/minute")
def get_client_timeline(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("opportunity_events")
        .select("event_type, event_description, created_at")
        .eq("client_id", client_id)
        .eq("company_id", tenant.company_id)
        .order("created_at", desc=True)
        .execute()
    )
    return res.data


@app.get("/api/pipeline")
@limiter.limit("30/minute")
def pipeline(request: Request, tenant: AuthTenant):
    db = admin_client if tenant.user_id == "engine_service" else scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, close_probability, score, priority_score, status, notes, estimated_annual_export_kwh, estimated_battery_savings, battery_payback_years, battery_opportunity_score, sales_script_long, sales_script_short, opportunity_reason")
        .eq("company_id", tenant.company_id)
        .gte("score", 40)
        .order("priority_score", desc=True)
        .limit(20)
        .execute()
    )
    data = res.data or []

    for c in data:
        slug = c.get("opportunity_type")
        c["opportunity_type_display"] = opportunity_display_es(slug)
        
    return data


@app.get("/api/opportunities")
@limiter.limit("30/minute")
def opportunities(request: Request, tenant: AuthTenant, limit: int = 100):
    db = admin_client if tenant.user_id == "engine_service" else scoped_client(tenant.jwt)
    safe_limit = max(1, min(limit, 500))
    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, score, priority_score, status")
        .eq("company_id", tenant.company_id)
        .gte("score", 40)
        .order("priority_score", desc=True)
        .limit(safe_limit)
        .execute()
    )
    data = res.data or []
    for item in data:
        item["opportunity_type_display"] = opportunity_display_es(item.get("opportunity_type"))
    return data


@app.get("/api/revenue-recovery")
@limiter.limit("20/minute")
def revenue_recovery(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, portal_invited, portal_leads(id)")
        .eq("company_id", tenant.company_id)
        .eq("portal_invited", True)
        .eq("portal_enabled", True)
        .execute()
    )
    
    recovery_clients = [
        client for client in (res.data or [])
        if not client.get("portal_leads")
    ]
    
    # Strip the relational data to keep the JSON payload minimal back to the client
    for client in recovery_clients:
        client.pop("portal_leads", None)
        
    return recovery_clients


@app.get("/api/opportunity-performance")
@limiter.limit("20/minute")
def opportunity_performance(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("opportunity_type, portal_invited, portal_leads(id)")
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not res.data:
        return []

    # Map aggregations
    performance_map = {}
    
    for c in res.data:
        opp_type = c.get("opportunity_type")
        if not opp_type:
            continue
            
        if opp_type not in performance_map:
            performance_map[opp_type] = {"detected_count": 0, "lead_count": 0}
            
        performance_map[opp_type]["detected_count"] += 1
        
        # Count explicit generated leads connected back from the unauthenticated portal
        leads = c.get("portal_leads") or []
        performance_map[opp_type]["lead_count"] += len(leads)
        
    # Final array mapping
    result = []
    for opp_type, data in performance_map.items():
        detected = data["detected_count"]
        leads = data["lead_count"]
        rate = leads / detected if detected > 0 else 0.0
        
        result.append({
            "opportunity_type": opp_type,
            "detected_count": detected,
            "lead_count": leads,
            "conversion_rate": round(rate, 2)
        })
        
    return sorted(result, key=lambda x: x["lead_count"], reverse=True)


@app.get("/api/portal-leads")
@limiter.limit("20/minute")
def get_portal_leads(request: Request, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    # We must join `portal_leads` matching over RLS to fetch recent notifications safely
    # Supabase allows native foreign-key selection syntax natively joining over the `client_id` parameter map.
    res = (
        db.table("portal_leads")
        .select("id, interest_type, requested_at, status, clients!inner(client_alias, company_id)")
        .eq("clients.company_id", tenant.company_id)
        .order("requested_at", desc=True)
        .limit(10)
        .execute()
    )
    
    if not res.data:
        return []
        
    leads = []
    for lead in res.data:
        client_data = lead.get("clients", {})
        status_raw = lead.get("status")
        leads.append({
            "id": lead.get("id"),
            "client_alias": client_data.get("client_alias", "Cliente"),
            "interest_type": lead.get("interest_type"),
            "requested_at": lead.get("requested_at"),
            "status": status_raw,
            "status_display": "Nuevo lead desde portal" if status_raw == "New" else status_raw
        })
        
    return leads


@app.get("/api/client/{client_id}/portal-analytics")
@limiter.limit("30/minute")
def get_portal_analytics(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    # Verify client ownership
    check = db.table("clients").select("id").eq("id", client_id).eq("company_id", tenant.company_id).single().execute()
    if not check.data:
        raise HTTPException(status_code=404, detail="Client not found")
        
    events_res = (
        db.table("portal_events")
        .select("event_type, created_at")
        .eq("client_id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    events = events_res.data or []
    
    portal_views = sum(1 for e in events if e["event_type"] == "portal_opened")
    proposal_downloads = sum(1 for e in events if e["event_type"] == "proposal_downloaded")
    consultation_requests = sum(1 for e in events if e["event_type"] == "consultation_requested")
    
    last_viewed = None
    view_events = [e for e in events if e["event_type"] == "portal_opened"]
    if view_events:
        last_viewed = max(v["created_at"] for v in view_events)
        
    return {
        "portal_views": portal_views,
        "proposal_downloads": proposal_downloads,
        "consultation_requests": consultation_requests,
        "last_viewed_at": last_viewed
    }


@app.get("/api/client/{client_id}")
@limiter.limit("30/minute")
def get_client(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = db.table("clients").select("*").eq("id", client_id).eq("company_id", tenant.company_id).single().execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found")
    return res.data


@app.post("/api/client/{client_id}/contacted")
@limiter.limit("30/minute")
def client_contacted(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    # Needs to check current status to potentially upgrade New -> Contacted
    status_res = db.table("clients").select("status").eq("id", client_id).eq("company_id", tenant.company_id).execute()
    
    if not status_res.data:
        raise HTTPException(status_code=404, detail="Client not found or access denied")
        
    current_status = status_res.data[0].get("status")
    now_ts = datetime.now(timezone.utc).isoformat()
    
    update_data = {
        "last_contact_at": now_ts
    }
    
    if current_status == "New":
        update_data["status"] = "Contacted"
        update_data["status_updated_at"] = now_ts
        
    res = (
        db.table("clients")
        .update(update_data)
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    return res.data[0]


# Moved above

@app.post("/api/client/{client_id}/notes")
@limiter.limit("30/minute")
def update_client_notes(request: Request, client_id: str, payload: NotesUpdatePayload, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    update_data = {
        "notes": payload.notes,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    res = (
        db.table("clients")
        .update(update_data)
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found or update failed")
        
    return res.data[0]


@app.get("/api/client/{client_id}/email-draft")
@limiter.limit("30/minute")
def get_client_email_draft(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    res = (
        db.table("clients")
        .select("id, client_alias, opportunity_type, expected_value, estimated_battery_savings, battery_payback_years, sales_script_short")
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .single()
        .execute()
    )

    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found")

    return build_sales_email_draft(res.data)


# Moved above


# Moved above


@app.post("/api/client/{client_id}/sales-action")
@limiter.limit("30/minute")
def log_sales_action(request: Request, client_id: str, payload: SalesActionPayload, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    now_ts = datetime.now(timezone.utc).isoformat()

    client_res = (
        db.table("clients")
        .select("id, client_alias, status")
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .single()
        .execute()
    )
    if not client_res.data:
        raise HTTPException(status_code=404, detail="Client not found")

    action_slug = payload.action.value
    action_label = SALES_ACTION_DISPLAY_ES.get(action_slug, action_slug)
    event_description = action_label
    if payload.note.strip():
        event_description = f"{action_label}. Nota: {payload.note.strip()}"

    try:
        event_res = db.table("opportunity_events").insert({
            "client_id": client_id,
            "company_id": tenant.company_id,
            "event_type": action_slug,
            "event_description": event_description,
        }).execute()
    except Exception as e:
        logger.error(f"Sales action insert failed: {e}")
        raise HTTPException(status_code=500, detail="No se pudo guardar la acción comercial.")

    current_status = client_res.data.get("status")
    update_data = {
        "last_contact_at": now_ts,
        "updated_at": now_ts,
    }
    if action_slug == "proposal_sent":
        if current_status in ("New", "Contacted"):
            update_data["status"] = "Proposal"
            update_data["status_updated_at"] = now_ts
    elif action_slug in ("called", "email_sent"):
        if current_status == "New":
            update_data["status"] = "Contacted"
            update_data["status_updated_at"] = now_ts

    db.table("clients").update(update_data).eq("id", client_id).eq("company_id", tenant.company_id).execute()

    event_row = event_res.data[0] if event_res.data else None
    return {
        "success": True,
        "message": "Acción comercial registrada",
        "event": event_row,
    }


# ─── Endpoints: Client Portal Foundation (BLOCK F8) ────────────────────────

@app.post("/api/client/{client_id}/portal-enable")
@limiter.limit("10/minute")
def enable_client_portal(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    token = secrets.token_urlsafe(32)
    
    res = (
        db.table("clients")
        .update({
            "portal_enabled": True,
            "portal_token": token,
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Client not found or update failed")
        
    return {"portal_url": f"/portal/{token}"}


@app.post("/api/client/{client_id}/send-portal")
@limiter.limit("10/minute")
def send_client_portal(request: Request, client_id: str, tenant: Tenant):
    db = scoped_client(tenant.jwt)
    
    client_res = db.table("clients").select("portal_token").eq("id", client_id).eq("company_id", tenant.company_id).single().execute()
    
    if not client_res.data:
        raise HTTPException(status_code=404, detail="Client not found or access denied")
        
    token = client_res.data.get("portal_token")
    if not token:
        token = secrets.token_urlsafe(32)
        
    update_data = {
        "portal_enabled": True,
        "portal_invited": True,
        "portal_token": token,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    update_res = (
        db.table("clients")
        .update(update_data)
        .eq("id", client_id)
        .eq("company_id", tenant.company_id)
        .execute()
    )
    
    if not update_res.data:
        raise HTTPException(status_code=500, detail="Failed to generate portal")
        
    return {"portal_url": f"/portal/{token}"}


@app.get("/api/public/portal/{token}")
@limiter.limit("30/minute")
def get_public_portal(request: Request, token: str):
    res = (
        admin_client.table("clients")
        .select("id, company_id, client_alias, system_size_kwp, installation_year, location_type, opportunity_type, expected_value")
        .eq("portal_token", token)
        .eq("portal_enabled", True)
        .single()
        .execute()
    )
    
    if not res.data:
        raise HTTPException(status_code=404, detail="Portal not found or inactive")
        
    # Log event: portal_opened
    try:
        admin_client.table("portal_events").insert({
            "client_id": res.data["id"],
            "company_id": res.data["company_id"],
            "event_type": "portal_opened",
            "event_metadata": {"ip": request.client.host if request.client else "unknown"}
        }).execute()
    except Exception as e:
        logger.warning("Failed to log portal_opened event: %s", str(e))
        
    return res.data


@app.post("/api/public/portal/{token}/request-proposal")
@limiter.limit("5/minute")
def request_portal_proposal(request: Request, token: str):
    res_client = (
        admin_client.table("clients")
        .select("id, company_id, opportunity_type")
        .eq("portal_token", token)
        .eq("portal_enabled", True)
        .single()
        .execute()
    )
    
    if not res_client.data:
        raise HTTPException(status_code=404, detail="Portal not found or inactive")
        
    client_data = res_client.data
    
    # Check if this lead was recently requested to prevent spam
    threshold_date = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    existing_lead = (
        admin_client.table("portal_leads")
        .select("id")
        .eq("client_id", client_data["id"])
        .gt("requested_at", threshold_date)
        .limit(1)
        .execute()
    )
    
    if existing_lead.data:
        logger.info("Duplicate portal lead prevented")
        return {"success": True}
        
    res_lead = (
        admin_client.table("portal_leads")
        .insert({
            "client_id": client_data["id"],
            "portal_token": token,
            "interest_type": client_data["opportunity_type"],
            "status": "New"
        })
        .execute()
    )
    
    # Log event: consultation_requested
    try:
        admin_client.table("portal_events").insert({
            "client_id": client_data["id"],
            "company_id": res_client.data["company_id"] if "company_id" in res_client.data else (admin_client.table("clients").select("company_id").eq("id", client_data["id"]).single().execute().data["company_id"]),
            "event_type": "consultation_requested"
        }).execute()
    except Exception as e:
        logger.warning("Failed to log consultation_requested event: %s", str(e))

    # Feedback loop visible in installer timeline/dashboard
    try:
        admin_client.table("opportunity_events").insert({
            "client_id": client_data["id"],
            "company_id": client_data["company_id"],
            "event_type": "portal_lead_received",
            "event_description": "Nuevo lead desde portal"
        }).execute()
    except Exception as e:
        logger.error(f"Failed to log portal lead feedback event: {e}")
        
    # ─── Step 9: Notify Installer via Email (Resend) ──────────────────────────
    if ENVIRONMENT == "production" and RESEND_API_KEY and RESEND_API_KEY != "re_test_placeholder":
        try:
            # 1. Fetch company email
            comp_res = admin_client.table("companies").select("email, name").eq("id", client_data["company_id"]).single().execute()
            if comp_res.data and comp_res.data.get("email"):
                comp_email = comp_res.data["email"]
                comp_name = comp_res.data.get("name", "Instalador")
                
                # 2. Fetch client details for the email
                c_res = admin_client.table("clients").select("client_alias").eq("id", client_data["id"]).single().execute()
                c_alias = c_res.data.get("client_alias", "Un cliente")
                
                # 3. Send email using Resend
                resend.Emails.send({
                    "from": "Solvist Notifications <notifications@solvist.io>",
                    "to": [comp_email],
                    "subject": f"NUEVO LEAD: Solicitud de propuesta de {c_alias}",
                    "html": f"""
                        <div style="font-family: sans-serif; padding: 20px; color: #1f2937;">
                            <h1 style="color: #2563eb;">Nuevo Lead Detectado</h1>
                            <p>Hola {comp_name},</p>
                            <p>El cliente <strong>{c_alias}</strong> ha revisado su portal y ha solicitado una propuesta comercial para:</p>
                            <div style="background: #f3f4f6; padding: 15px; border-radius: 8px; margin: 15px 0;">
                                <strong>{opportunity_display_es(client_data['opportunity_type'])}</strong>
                            </div>
                            <p>Puedes ver los detalles y gestionar este cliente en tu dashboard de Solvist.</p>
                            <hr style="border: 0; border-top: 1px solid #e5e7eb; margin: 20px 0;" />
                            <p style="font-size: 12px; color: #6b7280;">Este es un mensaje automático de Solvist Opportunity Intelligence.</p>
                        </div>
                    """
                })
                logger.info(f"Notification email sent to {comp_email} for lead from {c_alias}")
        except Exception as email_err:
            logger.error(f"Failed to send proposal notification email: {email_err}")

    return {"success": True}


# ─── Endpoints: System Diagnostics & Health (BLOCK R5) ────────────────────────

@app.get("/health")
def health_check():
    return {"status": "ok"}


@app.get("/api/import/status")
@limiter.limit("30/minute")
def import_status(request: Request, tenant: AuthTenant):
    return {
        "last_import_rows": int(LAST_IMPORT_STATUS.get("last_import_rows", 0)),
        "last_import_duration": float(LAST_IMPORT_STATUS.get("last_import_duration", 0.0)),
        "last_import_opportunities": int(LAST_IMPORT_STATUS.get("last_import_opportunities", 0)),
    }


@app.get("/api/auth/status")
@limiter.limit("30/minute")
def auth_status(request: Request, current_user: CurrentUser):
    return {
        "user_id": current_user.id,
        "status": "authenticated",
    }


@app.get("/internal/metrics/csv", include_in_schema=False)
@limiter.limit("30/minute")
def internal_csv_metrics(request: Request, tenant: InternalMetricsTenant):
    logger.info("CSV metrics accessed")
    return _get_csv_ingestion_metrics_snapshot()


@app.get("/health/pipeline")
@limiter.limit("30/minute")
def health_pipeline(request: Request, tenant: AuthTenant):
    database_status = "ok"
    try:
        admin_client.table("companies").select("id").limit(1).execute()
    except Exception:
        database_status = "error"

    auth_status = "ok" if _normalize_secret(os.getenv("ENGINE_SECRET")) else "error"
    scoring_status = "ok" if callable(core_score_all_installations) else "error"
    scheduler_status = "running" if getattr(scheduler, "running", False) else "stopped"

    return {
        "database": database_status,
        "supabase": database_status,
        "auth": auth_status,
        "scoring_engine": scoring_status,
        "scheduler": scheduler_status,
    }


@app.get("/version")
@limiter.limit("60/minute")
def version(request: Request):
    return {"version": DEPLOY_VERSION}


@app.get("/api/version")
@limiter.limit("60/minute")
def api_version(request: Request):
    return {"version": DEPLOY_VERSION}


@app.get("/api/system-check")
@limiter.limit("10/minute")
def system_check(request: Request):
    if ENVIRONMENT == "production":
        raise HTTPException(status_code=404, detail="Not Found")

    # Uses service role strictly for top-level diagnostic counts
    clients_res = admin_client.table("clients").select("id", count="exact").execute()
    clients_count = clients_res.count if hasattr(clients_res, "count") and clients_res.count is not None else len(clients_res.data)
    
    comp_res = admin_client.table("companies").select("id", count="exact").execute()
    companies_count = comp_res.count if hasattr(comp_res, "count") and comp_res.count is not None else len(comp_res.data)
    
    users_res = admin_client.table("users").select("id", count="exact").execute()
    users_count = users_res.count if hasattr(users_res, "count") and users_res.count is not None else len(users_res.data)
    
    inst_res = admin_client.table("installations").select("id", count="exact").execute()
    installations_count = inst_res.count if hasattr(inst_res, "count") and inst_res.count is not None else len(inst_res.data)
    
    return {
        "clients_count": clients_count,
        "companies_count": companies_count,
        "users_count": users_count,
        "installations_count": installations_count
    }


@app.get("/db-test")
@limiter.limit("10/minute")
def db_test(request: Request):
    if ENVIRONMENT == "production":
        raise HTTPException(status_code=404, detail="Not Found")

    try:
        conn = get_db_connection()
    except Exception:
        raise HTTPException(status_code=500, detail="Database connection not configured.")

    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM public.companies ORDER BY created_at ASC LIMIT 1;")
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=400, detail="No company found to attach installation.")
                company_id = row[0]

                cur.execute(
                    """
                    INSERT INTO public.installations (
                        company_id,
                        installation_year,
                        kwp,
                        inverter_model,
                        has_battery,
                        location_type,
                        country
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id;
                    """,
                    (company_id, 2020, 5.0, "DB Test Inverter", False, "residential", "ES"),
                )
                inserted_id = cur.fetchone()[0]
        return {"inserted_id": str(inserted_id)}
    except HTTPException:
        raise
    except psycopg2.Error:
        raise HTTPException(status_code=500, detail="Database insert failed.")
    finally:
        conn.close()
